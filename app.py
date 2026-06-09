import os
import io
import sqlite3
import datetime
import secrets
import hashlib
import json
import shutil
import zipfile
import string
import time
from functools import wraps

# ── TIMEZONE: ตั้งเป็นเวลาประเทศไทย (UTC+7) ──
# จำเป็นบน cloud server (PythonAnywhere/Render ใช้ UTC เป็น default)
os.environ['TZ'] = 'Asia/Bangkok'
try:
    time.tzset()  # POSIX only
except AttributeError:
    pass

from flask import Flask, request, jsonify, send_from_directory, send_file, session, redirect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── TIMEZONE ──────────────────────────────────────────────
# Force Asia/Bangkok for both Python datetime and SQLite's 'localtime' modifier.
# PythonAnywhere servers default to UTC, so without this all timestamps are -7 hours off.
os.environ['TZ'] = 'Asia/Bangkok'
try:
    time.tzset()  # POSIX (Linux/Mac) — not available on Windows
except AttributeError:
    pass

BASE = os.path.dirname(os.path.abspath(__file__))

# Persistent data location (use mounted disk on cloud, local 'data/' otherwise)
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(BASE, 'data'))
DB_PATH = os.path.join(DATA_DIR, 'school.db')
SECRET_FILE = os.path.join(DATA_DIR, '.secret_key')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(os.path.join(BASE, 'uploads'), exist_ok=True)

# Photos & assets: store under persistent dir on cloud so they survive redeploys
PHOTOS_DIR = os.path.join(DATA_DIR, 'photos')
ASSETS_DIR = os.path.join(DATA_DIR, 'assets')
os.makedirs(PHOTOS_DIR, exist_ok=True)
os.makedirs(ASSETS_DIR, exist_ok=True)

# Mirror to public/ via symlink so existing /photos/* and /assets/* URLs keep working
def _symlink_to_public(src, dst):
    try:
        if os.path.islink(dst):
            if os.path.realpath(dst) == os.path.realpath(src): return
            os.remove(dst)
        elif os.path.isdir(dst):
            # Existing dir with content: migrate files to DATA_DIR then replace with symlink
            import shutil
            for f in os.listdir(dst):
                s, d = os.path.join(dst, f), os.path.join(src, f)
                if not os.path.exists(d):
                    try: shutil.move(s, d)
                    except OSError: pass
            try: shutil.rmtree(dst)
            except OSError: return
        elif os.path.exists(dst):
            os.remove(dst)
        os.symlink(src, dst)
    except OSError: pass

_symlink_to_public(PHOTOS_DIR, os.path.join(BASE, 'public', 'photos'))
_symlink_to_public(ASSETS_DIR, os.path.join(BASE, 'public', 'assets'))

# Persistent secret key
if not os.path.exists(SECRET_FILE):
    with open(SECRET_FILE, 'w') as f:
        f.write(secrets.token_hex(32))
with open(SECRET_FILE) as f:
    SECRET_KEY = f.read().strip()

app = Flask(__name__, static_folder='public', static_url_path='')
app.secret_key = SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=7)
# Cache static files for 1 day (CSS/JS/images) — ลดโหลด CPU + Bandwidth
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 86400

@app.after_request
def add_cache_headers(response):
    """Cache static assets; HTML/API responses ไม่ cache"""
    path = request.path or ''
    # Service worker — ต้องไม่ cache (update เร็ว)
    if path == '/sw.js':
        response.headers['Cache-Control'] = 'no-cache, must-revalidate'
    # Manifest — cache 1 ชม.
    elif path == '/manifest.json':
        response.headers['Cache-Control'] = 'public, max-age=3600'
    # API + HTML — ไม่ cache
    elif path.startswith('/api/') or path.endswith('.html') or path == '/':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
    # Static assets — cache 1 วัน
    elif (path.startswith('/css/') or path.startswith('/js/')
          or path.startswith('/photos/') or path.startswith('/logos/')
          or path.startswith('/icons/')
          or path.endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg', '.ico', '.webp'))):
        response.headers['Cache-Control'] = 'public, max-age=86400'
    return response

# ── DATABASE ──────────────────────────────────────────────

def get_db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute('PRAGMA journal_mode=WAL')
    con.execute('PRAGMA foreign_keys=ON')
    return con

def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100_000)
    return f"{salt}${h.hex()}"

def verify_password(password, hashed):
    try:
        salt, _ = hashed.split('$', 1)
        return hash_password(password, salt) == hashed
    except Exception:
        return False

DEFAULT_SETTINGS = {
    'start_score':       '100',
    'deduct_absent':     '2',
    'deduct_late':       '1',
    'deduct_leave':      '0',
    'makeup_late_points':'1',
    'alert_threshold':   '5',
    'sem1_start':        '05-16',
    'sem1_end':          '10-15',
    'sem2_start':        '11-01',
    'sem2_end':          '03-31',
    'school_name':       'โรงเรียนตะกั่วทุ่งงานทวีวิทยาคม',
    'theme_color':       '#1a5276',
    'school_logo':       '',
    'director_name':     'นางสาวปัทมา จีนดี',
    'director_title':    'ผู้อำนวยการ',
    'director_view_key': '',   # secret key สำหรับเข้าหน้า ผอ. (auto-gen ตอน init)
}

def init_db():
    with get_db() as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                username        TEXT UNIQUE NOT NULL,
                password_hash   TEXT NOT NULL,
                full_name       TEXT NOT NULL,
                role            TEXT NOT NULL DEFAULT 'teacher'
                                CHECK (role IN ('admin','teacher')),
                assigned_level  INTEGER,
                assigned_room   INTEGER,
                must_change_pw  INTEGER DEFAULT 0,
                created_at      TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS login_attempts (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                username   TEXT,
                ip         TEXT,
                success    INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_login_created ON login_attempts(created_at);

            CREATE TABLE IF NOT EXISTS students (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                number       INTEGER,
                student_code TEXT,
                name         TEXT NOT NULL,
                class_level  INTEGER NOT NULL CHECK (class_level BETWEEN 1 AND 6),
                room         INTEGER NOT NULL,
                gender       TEXT,
                created_at   TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS attendance (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id  INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
                date        TEXT NOT NULL,
                status      TEXT NOT NULL DEFAULT 'present'
                                CHECK (status IN ('present','absent','late','leave','activity')),
                note        TEXT,
                recorded_by INTEGER REFERENCES users(id),
                recorded_at TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(student_id, date)
            );

            CREATE TABLE IF NOT EXISTS behavior_logs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id  INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
                date        TEXT NOT NULL,
                points      INTEGER NOT NULL,
                reason      TEXT NOT NULL,
                source      TEXT DEFAULT 'manual'
                                CHECK (source IN ('manual','attendance','makeup','uniform')),
                source_id   INTEGER,
                note        TEXT,
                recorded_by INTEGER REFERENCES users(id),
                created_at  TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS behavior_rules (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                points     INTEGER NOT NULL,
                category   TEXT,
                active     INTEGER DEFAULT 1,
                sort_order INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS pc_forms (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                form_type   TEXT NOT NULL,
                student_id  INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
                title       TEXT,
                payload     TEXT,
                recorded_by INTEGER REFERENCES users(id),
                created_at  TEXT DEFAULT (datetime('now','localtime')),
                updated_at  TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_pcf_student ON pc_forms(student_id);
            CREATE INDEX IF NOT EXISTS idx_pcf_type    ON pc_forms(form_type);

            CREATE TABLE IF NOT EXISTS home_visits (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id    INTEGER NOT NULL UNIQUE REFERENCES students(id) ON DELETE CASCADE,
                visited       INTEGER DEFAULT 0,         -- เยี่ยมแล้ว=1
                visit_date    TEXT,
                data          TEXT,                       -- JSON: ที่อยู่/ผู้ปกครอง/สภาพบ้าน/เศรษฐกิจ ฯลฯ
                photos        TEXT,                       -- JSON array ของชื่อไฟล์รูป
                fill_code     TEXT,                       -- รหัสลิงก์ให้นักเรียนกรอกเอง
                submitted     INTEGER DEFAULT 0,          -- นักเรียนกรอก+ส่งแล้ว=1 (รอครูยืนยัน)
                confirmed     INTEGER DEFAULT 0,          -- ครูยืนยันแล้ว=1
                recorded_by   INTEGER REFERENCES users(id),
                created_at    TEXT DEFAULT (datetime('now','localtime')),
                updated_at    TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_hv_student ON home_visits(student_id);

            CREATE TABLE IF NOT EXISTS holidays (
                date       TEXT PRIMARY KEY,
                name       TEXT NOT NULL,
                type       TEXT DEFAULT 'holiday'
                              CHECK (type IN ('holiday','event','exam')),
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS audit_logs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER,
                user_name  TEXT,
                action     TEXT NOT NULL,
                entity     TEXT,
                entity_id  INTEGER,
                detail     TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_logs(created_at);
            CREATE INDEX IF NOT EXISTS idx_audit_user    ON audit_logs(user_id);

            CREATE INDEX IF NOT EXISTS idx_att_date         ON attendance(date);
            CREATE INDEX IF NOT EXISTS idx_att_student      ON attendance(student_id);
            CREATE INDEX IF NOT EXISTS idx_att_student_date ON attendance(student_id, date);
            CREATE INDEX IF NOT EXISTS idx_att_date_status  ON attendance(date, status);
            CREATE INDEX IF NOT EXISTS idx_stu_class        ON students(class_level, room);
            CREATE INDEX IF NOT EXISTS idx_bhv_student      ON behavior_logs(student_id);
            CREATE INDEX IF NOT EXISTS idx_bhv_source       ON behavior_logs(source, source_id);
            CREATE INDEX IF NOT EXISTS idx_bhv_date         ON behavior_logs(date);
            CREATE INDEX IF NOT EXISTS idx_bhv_student_date ON behavior_logs(student_id, date);
        """)
        con.execute("ANALYZE")  # update query planner statistics

        # Migrate: users
        cols = [r['name'] for r in con.execute('PRAGMA table_info(users)').fetchall()]
        if 'assigned_level' not in cols:
            con.execute('ALTER TABLE users ADD COLUMN assigned_level INTEGER')
        if 'assigned_room' not in cols:
            con.execute('ALTER TABLE users ADD COLUMN assigned_room INTEGER')
        if 'must_change_pw' not in cols:
            con.execute('ALTER TABLE users ADD COLUMN must_change_pw INTEGER DEFAULT 0')
        # Migrate: students - add photo + parent_code + birthdate + national_id
        scols = [r['name'] for r in con.execute('PRAGMA table_info(students)').fetchall()]
        if 'photo' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN photo TEXT')
        if 'parent_code' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN parent_code TEXT')
            con.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_stu_pcode ON students(parent_code)')
        if 'birthdate' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN birthdate TEXT')  # ISO YYYY-MM-DD
        if 'national_id' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN national_id TEXT')  # 13 หลัก
        # ข้อมูลผู้ปกครอง (สำหรับ ปค.3 ทัณฑ์บน) — กรอกเฉพาะรายที่มีเคส
        if 'parent_name' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN parent_name TEXT')
        if 'parent_relation' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN parent_relation TEXT')  # บิดา/มารดา/...
        if 'parent_phone' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN parent_phone TEXT')
        if 'address' not in scols:
            con.execute('ALTER TABLE students ADD COLUMN address TEXT')

        # Migrate: home_visits - add fill_code/submitted/confirmed (สำหรับลิงก์ให้นักเรียนกรอก)
        hvcols = [r['name'] for r in con.execute('PRAGMA table_info(home_visits)').fetchall()]
        if hvcols:  # ตารางมีอยู่แล้ว
            if 'fill_code' not in hvcols:
                con.execute('ALTER TABLE home_visits ADD COLUMN fill_code TEXT')
                con.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_hv_fillcode ON home_visits(fill_code)')
            if 'submitted' not in hvcols:
                con.execute('ALTER TABLE home_visits ADD COLUMN submitted INTEGER DEFAULT 0')
            if 'confirmed' not in hvcols:
                con.execute('ALTER TABLE home_visits ADD COLUMN confirmed INTEGER DEFAULT 0')

        # Migrate: behavior_logs - add note column
        bcols = [r['name'] for r in con.execute('PRAGMA table_info(behavior_logs)').fetchall()]
        if 'note' not in bcols:
            con.execute('ALTER TABLE behavior_logs ADD COLUMN note TEXT')

        # Migrate: attendance - allow 'activity' status (rebuild if old CHECK constraint)
        try:
            con.execute("INSERT INTO attendance (student_id, date, status) VALUES (-9999, '0000-01-01', 'activity')")
            con.execute("DELETE FROM attendance WHERE student_id=-9999")
        except sqlite3.IntegrityError as e:
            if 'CHECK' in str(e) or 'check' in str(e).lower():
                con.executescript("""
                    CREATE TABLE attendance_new (
                        id          INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id  INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
                        date        TEXT NOT NULL,
                        status      TEXT NOT NULL DEFAULT 'present'
                                        CHECK (status IN ('present','absent','late','leave','activity')),
                        note        TEXT,
                        recorded_by INTEGER REFERENCES users(id),
                        recorded_at TEXT DEFAULT (datetime('now','localtime')),
                        UNIQUE(student_id, date)
                    );
                    INSERT INTO attendance_new (id, student_id, date, status, note, recorded_by, recorded_at)
                      SELECT id, student_id, date, status, note, recorded_by, recorded_at FROM attendance;
                    DROP TABLE attendance;
                    ALTER TABLE attendance_new RENAME TO attendance;
                    CREATE INDEX IF NOT EXISTS idx_att_date    ON attendance(date);
                    CREATE INDEX IF NOT EXISTS idx_att_student ON attendance(student_id);
                """)
                print("  [Migrate] อัพเดทตาราง attendance รองรับสถานะ 'activity'")

        # Migrate: behavior_logs.source - allow 'makeup' + 'uniform'
        try:
            con.execute("INSERT INTO behavior_logs (student_id, date, points, reason, source) VALUES (-9999, '0000-01-01', 0, '_t', 'uniform')")
            con.execute("DELETE FROM behavior_logs WHERE student_id=-9999")
        except sqlite3.IntegrityError as e:
            if 'CHECK' in str(e) or 'check' in str(e).lower():
                con.executescript("""
                    CREATE TABLE behavior_logs_new (
                        id          INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id  INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
                        date        TEXT NOT NULL,
                        points      INTEGER NOT NULL,
                        reason      TEXT NOT NULL,
                        source      TEXT DEFAULT 'manual'
                                       CHECK (source IN ('manual','attendance','makeup','uniform')),
                        source_id   INTEGER,
                        note        TEXT,
                        recorded_by INTEGER REFERENCES users(id),
                        created_at  TEXT DEFAULT (datetime('now','localtime'))
                    );
                    INSERT INTO behavior_logs_new (id, student_id, date, points, reason, source, source_id, recorded_by, created_at)
                      SELECT id, student_id, date, points, reason, source, source_id, recorded_by, created_at FROM behavior_logs;
                    DROP TABLE behavior_logs;
                    ALTER TABLE behavior_logs_new RENAME TO behavior_logs;
                    CREATE INDEX IF NOT EXISTS idx_bhv_student ON behavior_logs(student_id);
                    CREATE INDEX IF NOT EXISTS idx_bhv_source  ON behavior_logs(source, source_id);
                """)
                print("  [Migrate] อัพเดท behavior_logs รองรับ source='uniform'")

        # Default settings
        for k, v in DEFAULT_SETTINGS.items():
            con.execute('INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)', (k, v))

        # Auto-gen director_view_key ถ้ายังว่าง
        cur = con.execute("SELECT value FROM settings WHERE key='director_view_key'").fetchone()
        if cur and not cur['value']:
            key = secrets.token_urlsafe(12)  # ~16 chars random
            con.execute("UPDATE settings SET value=? WHERE key='director_view_key'", (key,))
            print(f"  [Init] สร้าง director_view_key: {key}")

        # Default behavior rules — auto add missing entries (ใช้ name เป็น key ไม่ซ้ำ)
        defaults = [
            ('มาสาย',                    -1, 'การมาเรียน',     10),
            ('ขาดเรียนไม่มีเหตุผล',      -2, 'การมาเรียน',     20),
            ('หนีเรียน',                -10, 'การมาเรียน',     30),
            ('ไม่ส่งงาน/การบ้าน',        -3, 'การเรียน',       40),
            ('ทุจริตการสอบ',            -20, 'การเรียน',       50),
            ('ไม่ใส่เครื่องแบบให้ถูกต้อง', -5, 'การแต่งกาย',     60),
            ('ทรงผมผิดระเบียบ',          -5, 'การแต่งกาย',     65),
            ('เล็บยาว/ทาเล็บ',           -3, 'การแต่งกาย',     67),
            ('รองเท้า/ถุงเท้าผิดระเบียบ', -3, 'การแต่งกาย',     68),
            ('เครื่องประดับเกินจำเป็น',   -3, 'การแต่งกาย',     69),
            ('ใช้โทรศัพท์ในห้องเรียน',    -5, 'ระเบียบวินัย',   80),
            ('พูดจาไม่สุภาพ',            -5, 'ระเบียบวินัย',   90),
            ('ทะเลาะวิวาท',             -20, 'ระเบียบวินัย',  100),
            ('ทำร้ายร่างกายผู้อื่น',     -30, 'ระเบียบวินัย',  110),
            ('ทำลายทรัพย์สินโรงเรียน',  -20, 'ระเบียบวินัย',  120),
            ('สูบบุหรี่/บุหรี่ไฟฟ้า',   -30, 'สารเสพติด',     130),
            ('เสพ/ครอบครองสารเสพติด',   -50, 'สารเสพติด',     140),
            ('ลักทรัพย์',               -30, 'ระเบียบวินัย',  150),
            ('ช่วยเหลือกิจกรรมส่วนรวม',  +5, 'ความดี',        200),
            ('ทำคุณงามความดี',           +5, 'ความดี',        210),
        ]
        existing_names = {r['name'] for r in con.execute('SELECT name FROM behavior_rules').fetchall()}
        added = 0
        for name, points, category, sort_order in defaults:
            if name not in existing_names:
                con.execute(
                    'INSERT INTO behavior_rules (name, points, category, sort_order) VALUES (?,?,?,?)',
                    (name, points, category, sort_order)
                )
                added += 1
        if added:
            print(f'  [Init] เพิ่มกฎเริ่มต้น {added} รายการ')

        # Default admin account
        admin = con.execute('SELECT id FROM users WHERE username=?', ('admin',)).fetchone()
        if not admin:
            con.execute(
                'INSERT INTO users (username, password_hash, full_name, role, must_change_pw) VALUES (?, ?, ?, ?, 1)',
                ('admin', hash_password('admin123'), 'ผู้ดูแลระบบ', 'admin')
            )
            print('  [Init] สร้างบัญชี admin เริ่มต้น (admin/admin123) — ต้องเปลี่ยนรหัสตอน login ครั้งแรก')

init_db()

# ── HELPERS ───────────────────────────────────────────────

def rows_to_list(rows):
    return [dict(r) for r in rows]

def today_iso():
    return datetime.date.today().isoformat()

def get_settings():
    with get_db() as con:
        rows = con.execute('SELECT key, value FROM settings').fetchall()
    return {r['key']: r['value'] for r in rows}

def current_user():
    if 'user_id' not in session:
        return None
    with get_db() as con:
        u = con.execute("""SELECT id, username, full_name, role,
                                  assigned_level, assigned_room, must_change_pw
                           FROM users WHERE id=?""", (session['user_id'],)).fetchone()
    return dict(u) if u else None

def user_class_filter(user, args=None):
    """Return (extra_where, extra_params, effective_level, effective_room)
    suitable for queries on `students s`.

    - Admin: no restriction; honors ?level/?room from args if provided
    - Teacher with assigned class: forced to that class; ignores args
    - Teacher without assigned class: no restriction; honors ?level/?room
    """
    where, params = '', []
    eff_level = eff_room = None

    if user and user['role'] == 'teacher' and user.get('assigned_level'):
        where += ' AND s.class_level=?'; params.append(int(user['assigned_level']))
        eff_level = int(user['assigned_level'])
        if user.get('assigned_room'):
            where += ' AND s.room=?'; params.append(int(user['assigned_room']))
            eff_room = int(user['assigned_room'])
    elif args is not None:
        lv = args.get('level'); rm = args.get('room')
        if lv: where += ' AND s.class_level=?'; params.append(int(lv)); eff_level = int(lv)
        if rm: where += ' AND s.room=?';        params.append(int(rm)); eff_room = int(rm)

    return where, params, eff_level, eff_room

def audit_log(con, action, entity=None, entity_id=None, detail=None):
    """Record an action to the audit_logs table. Pass an open connection."""
    u = current_user()
    user_id = u['id'] if u else None
    user_name = u['full_name'] if u else None
    if isinstance(detail, (dict, list)):
        detail = json.dumps(detail, ensure_ascii=False)
    con.execute(
        'INSERT INTO audit_logs (user_id, user_name, action, entity, entity_id, detail) VALUES (?,?,?,?,?,?)',
        (user_id, user_name, action, entity, entity_id, detail)
    )

def generate_parent_code():
    """6-character alphanumeric code (no confusing chars)."""
    alphabet = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(secrets.choice(alphabet) for _ in range(6))

def can_access_student(user, student_row):
    """Check if user can access a particular student record."""
    if not user: return False
    if user['role'] == 'admin': return True
    if user.get('assigned_level') and student_row['class_level'] != user['assigned_level']:
        return False
    if user.get('assigned_room') and student_row['room'] != user['assigned_room']:
        return False
    return True

def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if 'user_id' not in session:
            return jsonify(error='unauthorized'), 401
        return f(*a, **kw)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if 'user_id' not in session:
            return jsonify(error='unauthorized'), 401
        u = current_user()
        if not u or u['role'] != 'admin':
            return jsonify(error='forbidden'), 403
        return f(*a, **kw)
    return wrapper

def apply_attendance_behavior(con, student_id, att_id, date, status, settings, user_id):
    """Recalculate behavior log for an attendance record. Idempotent."""
    # Remove old auto entry for this attendance
    con.execute('DELETE FROM behavior_logs WHERE source=? AND source_id=?', ('attendance', att_id))

    points_map = {
        'absent': -int(settings.get('deduct_absent', '2') or 0),
        'late':   -int(settings.get('deduct_late', '1') or 0),
        'leave':  -int(settings.get('deduct_leave', '0') or 0),
        # 'activity' และ 'present' ไม่หักคะแนน
    }
    reasons = {
        'absent': 'ขาดเรียน',
        'late':   'มาสาย',
        'leave':  'ลา',
    }
    if status in points_map and points_map[status] != 0:
        con.execute("""
            INSERT INTO behavior_logs (student_id, date, points, reason, source, source_id, recorded_by)
            VALUES (?, ?, ?, ?, 'attendance', ?, ?)
        """, (student_id, date, points_map[status], reasons[status], att_id, user_id))

# ── STATIC ────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect('/login.html')
    return send_from_directory('public', 'index.html')

# ── AUTH ──────────────────────────────────────────────────

MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

@app.post('/api/auth/login')
def api_login():
    body = request.get_json() or {}
    username = (body.get('username') or '').strip()
    password = body.get('password') or ''
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown').split(',')[0].strip()

    if not username or not password:
        return jsonify(success=False, message='กรุณากรอกชื่อผู้ใช้และรหัสผ่าน'), 400

    # Check rate limit (per IP + username combination)
    with get_db() as con:
        cutoff = (datetime.datetime.now() - datetime.timedelta(minutes=LOCKOUT_MINUTES)).strftime('%Y-%m-%d %H:%M:%S')
        recent_fails = con.execute(
            """SELECT COUNT(*) AS n FROM login_attempts
               WHERE (ip=? OR username=?) AND success=0 AND created_at > ?""",
            (ip, username, cutoff)
        ).fetchone()['n']

        if recent_fails >= MAX_LOGIN_ATTEMPTS:
            con.execute('INSERT INTO login_attempts (username, ip, success) VALUES (?,?,0)',
                        (username, ip))
            return jsonify(success=False,
                message=f'ล็อกอินผิดเกิน {MAX_LOGIN_ATTEMPTS} ครั้ง — กรุณารออีก {LOCKOUT_MINUTES} นาที'), 429

        u = con.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        ok = u and verify_password(password, u['password_hash'])

        con.execute('INSERT INTO login_attempts (username, ip, success) VALUES (?,?,?)',
                    (username, ip, 1 if ok else 0))

    if not ok:
        remaining = MAX_LOGIN_ATTEMPTS - recent_fails - 1
        msg = 'ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง'
        if remaining <= 2 and remaining >= 0:
            msg += f' (เหลืออีก {remaining} ครั้งก่อนถูกล็อก)'
        return jsonify(success=False, message=msg), 401

    session.permanent = True
    session['user_id'] = u['id']
    return jsonify(success=True, user={
        'id': u['id'], 'username': u['username'],
        'full_name': u['full_name'], 'role': u['role'],
        'assigned_level': u['assigned_level'], 'assigned_room': u['assigned_room'],
        'must_change_pw': bool(u['must_change_pw'])
    })

@app.post('/api/auth/logout')
def api_logout():
    session.clear()
    return jsonify(success=True)

@app.get('/api/auth/me')
def api_me():
    u = current_user()
    if not u:
        return jsonify(authenticated=False), 401
    return jsonify(authenticated=True, user=u)

@app.post('/api/auth/password')
@login_required
def api_change_password():
    body = request.get_json() or {}
    old = body.get('old', '')
    new = body.get('new', '')
    if len(new) < 6:
        return jsonify(success=False, message='รหัสผ่านใหม่ต้องยาวอย่างน้อย 6 ตัว'), 400
    if new == old:
        return jsonify(success=False, message='รหัสผ่านใหม่ต้องไม่เหมือนรหัสเดิม'), 400

    with get_db() as con:
        u = con.execute('SELECT password_hash FROM users WHERE id=?',
                        (session['user_id'],)).fetchone()
        if not verify_password(old, u['password_hash']):
            return jsonify(success=False, message='รหัสผ่านเดิมไม่ถูกต้อง'), 400
        con.execute('UPDATE users SET password_hash=?, must_change_pw=0 WHERE id=?',
                    (hash_password(new), session['user_id']))
        audit_log(con, 'change_own_password', 'user', session['user_id'], None)
    return jsonify(success=True, message='เปลี่ยนรหัสผ่านสำเร็จ')

# ── USER MANAGEMENT (admin) ───────────────────────────────

@app.get('/api/users')
@admin_required
def api_users_list():
    with get_db() as con:
        rows = con.execute("""SELECT id, username, full_name, role,
                                     assigned_level, assigned_room, created_at
                              FROM users ORDER BY username""").fetchall()
    return jsonify(rows_to_list(rows))

@app.post('/api/users')
@admin_required
def api_users_create():
    b = request.get_json() or {}
    username = (b.get('username') or '').strip()
    password = b.get('password') or ''
    full_name = (b.get('full_name') or '').strip()
    role = b.get('role', 'teacher')
    a_level = b.get('assigned_level')
    a_room  = b.get('assigned_room')
    a_level = int(a_level) if a_level else None
    a_room  = int(a_room)  if a_room  else None

    if not username or not password or not full_name:
        return jsonify(success=False, message='กรอกข้อมูลให้ครบ'), 400
    if role not in ('admin', 'teacher'):
        return jsonify(success=False, message='สิทธิ์ไม่ถูกต้อง'), 400
    if len(password) < 4:
        return jsonify(success=False, message='รหัสผ่านต้องยาวอย่างน้อย 4 ตัว'), 400
    if role == 'admin':
        a_level = a_room = None  # Admin doesn't have a class

    try:
        with get_db() as con:
            con.execute("""INSERT INTO users
                (username, password_hash, full_name, role, assigned_level, assigned_room, must_change_pw)
                VALUES (?,?,?,?,?,?,1)""",
                (username, hash_password(password), full_name, role, a_level, a_room))
            audit_log(con, 'create_user', 'user', None, {'username': username, 'role': role})
    except sqlite3.IntegrityError:
        return jsonify(success=False, message='ชื่อผู้ใช้นี้มีอยู่แล้ว'), 400

    return jsonify(success=True, message=f'สร้างบัญชี "{username}" สำเร็จ (ผู้ใช้ต้องเปลี่ยนรหัสผ่านตอน login ครั้งแรก)')

@app.put('/api/users/<int:uid>/assign')
@admin_required
def api_users_assign(uid):
    """Update assigned class for a teacher."""
    b = request.get_json() or {}
    a_level = b.get('assigned_level')
    a_room  = b.get('assigned_room')
    a_level = int(a_level) if a_level else None
    a_room  = int(a_room)  if a_room  else None
    with get_db() as con:
        u = con.execute('SELECT role FROM users WHERE id=?', (uid,)).fetchone()
        if not u: return jsonify(success=False, message='ไม่พบผู้ใช้'), 404
        if u['role'] == 'admin':
            return jsonify(success=False, message='แอดมินไม่ต้องกำหนดห้อง'), 400
        con.execute('UPDATE users SET assigned_level=?, assigned_room=? WHERE id=?',
                    (a_level, a_room, uid))
    return jsonify(success=True, message='กำหนดห้องเรียบร้อย')

@app.delete('/api/users/<int:uid>')
@admin_required
def api_users_delete(uid):
    if uid == session['user_id']:
        return jsonify(success=False, message='ลบบัญชีตัวเองไม่ได้'), 400
    with get_db() as con:
        con.execute('DELETE FROM users WHERE id=?', (uid,))
    return jsonify(success=True)

@app.post('/api/users/<int:uid>/reset-password')
@admin_required
def api_users_reset_password(uid):
    b = request.get_json() or {}
    new = b.get('password') or ''
    if len(new) < 4:
        return jsonify(success=False, message='รหัสผ่านต้องยาวอย่างน้อย 4 ตัว'), 400
    with get_db() as con:
        con.execute('UPDATE users SET password_hash=?, must_change_pw=1 WHERE id=?',
                    (hash_password(new), uid))
        audit_log(con, 'reset_password', 'user', uid, None)
    return jsonify(success=True, message='รีเซ็ตรหัสผ่านสำเร็จ (ผู้ใช้ต้องเปลี่ยนใหม่ตอน login ครั้งหน้า)')

# ── SETTINGS ──────────────────────────────────────────────

@app.get('/api/settings')
@login_required
def api_settings_get():
    return jsonify(get_settings())

@app.get('/api/public/theme')
def api_public_theme():
    """Public endpoint — returns only branding info (no auth required)."""
    s = get_settings()
    return jsonify(
        school_name=s.get('school_name'),
        school_logo=s.get('school_logo'),
        theme_color=s.get('theme_color'),
        director_name=s.get('director_name'),
        director_title=s.get('director_title'),
    )

@app.put('/api/settings')
@admin_required
def api_settings_update():
    b = request.get_json() or {}
    allowed = set(DEFAULT_SETTINGS.keys())
    # ป้องกัน director_view_key ถูกแก้ผ่าน UI (ต้องใช้ /director-key endpoint)
    allowed.discard('director_view_key')
    with get_db() as con:
        for k, v in b.items():
            if k in allowed:
                con.execute('INSERT INTO settings (key,value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',
                            (k, str(v)))
    return jsonify(success=True, message='บันทึกการตั้งค่าสำเร็จ')

@app.post('/api/settings/director-key')
@admin_required
def api_regen_director_key():
    """สร้าง director_view_key ใหม่ (ลิงก์เก่าจะใช้ไม่ได้)"""
    new_key = secrets.token_urlsafe(12)
    with get_db() as con:
        con.execute(
            'INSERT INTO settings (key,value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',
            ('director_view_key', new_key))
        audit_log(con, 'regen_director_key', 'settings', None, None)
    return jsonify(success=True, key=new_key, message='สร้าง key ใหม่เรียบร้อย')

# ── CLASSES ───────────────────────────────────────────────

@app.get('/api/classes')
@login_required
def api_classes():
    u = current_user()
    where, params, _, _ = user_class_filter(u)
    sql = f"""
        SELECT class_level, room, COUNT(*) as count
        FROM students s WHERE 1=1 {where}
        GROUP BY class_level, room ORDER BY class_level, room
    """
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))

# ── STUDENTS ─────────────────────────────────────────────

@app.get('/api/students')
@login_required
def api_students():
    u = current_user()
    where, params, _, _ = user_class_filter(u, request.args)
    # SELECT * already includes parent_code/photo via schema migration
    sql = f'SELECT s.* FROM students s WHERE 1=1 {where} ORDER BY class_level, room, number, name'
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))

@app.get('/api/students/template')
@login_required
def api_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'รายชื่อนักเรียน'

    headers = ['เลขที่', 'รหัสนักเรียน', 'ชื่อ-นามสกุล', 'ชั้น (1-6)', 'ห้อง', 'เพศ',
               'วันเกิด (YYYY-MM-DD)', 'เลขบัตรประชาชน (13 หลัก)']
    hfill = PatternFill(fill_type='solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    sample = [
        [1, '12345', 'เด็กชาย ตัวอย่าง นามสกุลตัวอย่าง', 1, 1, 'ชาย', '2012-05-15', '1234567890123'],
        [2, '12346', 'เด็กหญิง ตัวอย่าง นามสกุลตัวอย่าง', 1, 1, 'หญิง', '2012-08-22', ''],
        [3, '12347', 'เด็กชาย สมชาย รักเรียน', 2, 1, 'ชาย', '', ''],
    ]
    for r in sample: ws.append(r)
    for col, w in zip(ws.columns, [8, 16, 32, 12, 8, 8, 16, 20]):
        ws.column_dimensions[col[0].column_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='student_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.post('/api/students/import')
@admin_required
def api_import():
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify(success=False, message='ต้องเป็นไฟล์ Excel'), 400

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        return jsonify(success=False, message=f'เปิดไฟล์ไม่ได้: {e}'), 400

    count, errors, inserts = 0, 0, []
    for row in rows:
        if not row or len(row) < 3: continue
        try:
            num    = int(row[0]) if row[0] is not None else None
            code   = str(row[1]).strip() if row[1] is not None else None
            name   = str(row[2]).strip() if row[2] is not None else ''
            level  = int(row[3]) if row[3] is not None else None
            room   = int(row[4]) if row[4] is not None else None
            gender = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
            # วันเกิด (col 7) — รองรับทั้ง string และ datetime
            birthdate = None
            if len(row) > 6 and row[6] is not None:
                bd = row[6]
                if hasattr(bd, 'isoformat'):
                    birthdate = bd.date().isoformat() if hasattr(bd, 'date') else bd.isoformat()
                else:
                    bd_str = str(bd).strip()
                    if bd_str:
                        try:
                            datetime.date.fromisoformat(bd_str)
                            birthdate = bd_str
                        except ValueError: pass
            # เลขบัตรประชาชน (col 8) — must be 13 digits
            national_id = None
            if len(row) > 7 and row[7] is not None:
                nid = str(row[7]).strip().replace('-', '').replace(' ', '')
                if nid.isdigit() and len(nid) == 13:
                    national_id = nid
            if not name or not level or not room or not (1 <= level <= 6):
                errors += 1; continue
            inserts.append((num, code, name, level, room, gender, birthdate, national_id))
            count += 1
        except Exception:
            errors += 1

    if inserts:
        with get_db() as con:
            con.executemany(
                'INSERT INTO students (number,student_code,name,class_level,room,gender,birthdate,national_id) '
                'VALUES (?,?,?,?,?,?,?,?)',
                inserts
            )
            audit_log(con, 'import_students', 'student', None, {'count': count, 'errors': errors})

    msg = f'นำเข้าสำเร็จ {count} คน' + (f' (ข้าม {errors} แถว)' if errors else '')
    return jsonify(success=True, count=count, errors=errors, message=msg)

@app.get('/api/students/export')
@admin_required
def api_students_export():
    """Export รายชื่อนักเรียนปัจจุบันทั้งหมด เป็น Excel — ใช้ format เดียวกับ template
    เพื่อให้แก้แล้ว upload กลับผ่าน /bulk-update ได้ทันที"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'รายชื่อนักเรียน'

    headers = ['เลขที่', 'รหัสนักเรียน', 'ชื่อ-นามสกุล', 'ชั้น (1-6)', 'ห้อง', 'เพศ',
               'วันเกิด (YYYY-MM-DD)', 'เลขบัตรประชาชน (13 หลัก)']
    hfill = PatternFill(fill_type='solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    with get_db() as con:
        rows = con.execute("""
            SELECT number, student_code, name, class_level, room, gender, birthdate, national_id
            FROM students ORDER BY class_level, room, number, name
        """).fetchall()
    for r in rows:
        ws.append([r['number'], r['student_code'], r['name'], r['class_level'],
                   r['room'], r['gender'], r['birthdate'], r['national_id']])
    for col, w in zip(ws.columns, [8, 16, 32, 12, 8, 8, 16, 20]):
        ws.column_dimensions[col[0].column_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    ts = datetime.datetime.now().strftime('%Y%m%d')
    return send_file(buf, as_attachment=True,
                     download_name=f'students_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.post('/api/students/bulk-update')
@admin_required
def api_bulk_update():
    """อัพเดทข้อมูลนักเรียนเก่าจาก Excel — match ตาม student_code (priority) หรือ name+class+room
    คอลัมน์ที่ใช้ update: number, name, gender, birthdate, national_id (ยกเว้นช่องว่าง)
    """
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify(success=False, message='ต้องเป็นไฟล์ Excel'), 400

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        return jsonify(success=False, message=f'เปิดไฟล์ไม่ได้: {e}'), 400

    updated = 0
    not_found = []
    skipped = 0

    with get_db() as con:
        for row in rows:
            if not row or len(row) < 3: continue
            try:
                num   = int(row[0]) if row[0] is not None else None
                code  = str(row[1]).strip() if row[1] is not None else ''
                name  = str(row[2]).strip() if row[2] is not None else ''
                level = int(row[3]) if row[3] is not None else None
                room  = int(row[4]) if row[4] is not None else None
                gender = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None

                # birthdate
                birthdate = None
                if len(row) > 6 and row[6] is not None:
                    bd = row[6]
                    if hasattr(bd, 'isoformat'):
                        birthdate = bd.date().isoformat() if hasattr(bd, 'date') else bd.isoformat()
                    else:
                        bd_str = str(bd).strip()
                        if bd_str:
                            try:
                                datetime.date.fromisoformat(bd_str)
                                birthdate = bd_str
                            except ValueError: pass

                # national_id
                national_id = None
                if len(row) > 7 and row[7] is not None:
                    nid = str(row[7]).strip().replace('-', '').replace(' ', '')
                    if nid.isdigit() and len(nid) == 13:
                        national_id = nid

                # หานักเรียน — priority: student_code → name+class+room
                target = None
                if code:
                    target = con.execute(
                        'SELECT id FROM students WHERE student_code=?', (code,)
                    ).fetchone()
                if not target and name and level and room:
                    target = con.execute(
                        'SELECT id FROM students WHERE name=? AND class_level=? AND room=?',
                        (name, level, room)
                    ).fetchone()

                if not target:
                    not_found.append(name or code or f'row {num}')
                    continue

                # Build update — ใส่เฉพาะที่มีค่าใหม่ (ไม่ overwrite ของเดิมด้วยว่าง)
                updates = {}
                if num is not None: updates['number'] = num
                if name:            updates['name'] = name
                if gender:          updates['gender'] = gender
                if birthdate:       updates['birthdate'] = birthdate
                if national_id:     updates['national_id'] = national_id

                if not updates:
                    skipped += 1
                    continue

                cols = ', '.join(f'{k}=?' for k in updates.keys())
                con.execute(
                    f'UPDATE students SET {cols} WHERE id=?',
                    list(updates.values()) + [target['id']]
                )
                updated += 1

            except Exception:
                skipped += 1

        audit_log(con, 'bulk_update_students', 'student', None,
                  {'updated': updated, 'not_found': len(not_found), 'skipped': skipped})

    msg = f'อัพเดทสำเร็จ {updated} คน'
    if not_found: msg += f' • ไม่พบ {len(not_found)} คน'
    if skipped:   msg += f' • ข้าม {skipped} แถว'
    return jsonify(success=True, updated=updated, not_found=not_found, skipped=skipped, message=msg)

@app.delete('/api/students/all')
@admin_required
def api_students_clear():
    with get_db() as con:
        con.execute('DELETE FROM behavior_logs')
        con.execute('DELETE FROM attendance')
        con.execute('DELETE FROM students')
    return jsonify(success=True, message='ลบข้อมูลทั้งหมดแล้ว')

@app.get('/api/students/<int:sid>')
@login_required
def api_student_get(sid):
    """ข้อมูลนักเรียน 1 คน — สำหรับ confirm scan + อื่นๆ"""
    u = current_user()
    with get_db() as con:
        s = con.execute(
            'SELECT id, number, student_code, name, class_level, room, gender, photo, '
            'birthdate, national_id, parent_code, '
            'parent_name, parent_relation, parent_phone, address FROM students WHERE id=?',
            (sid,)).fetchone()
    if not s:
        return jsonify(error='ไม่พบนักเรียน'), 404
    # ครูดึงได้เฉพาะห้องตัวเอง
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if s['class_level'] != u['assigned_level']:
            return jsonify(error='ไม่มีสิทธิ์ดูข้อมูลห้องนี้'), 403
        if u.get('assigned_room') and s['room'] != u['assigned_room']:
            return jsonify(error='ไม่มีสิทธิ์ดูข้อมูลห้องนี้'), 403
    return jsonify(dict(s))

@app.get('/api/students/<int:sid>/basic')
@login_required
def api_student_basic(sid):
    """ข้อมูลพื้นฐานสำหรับ confirm ตอนสแกนมาสาย — ทุกคนดูได้ทุกห้อง (ไม่รวมข้อมูลอ่อนไหว)"""
    with get_db() as con:
        s = con.execute(
            'SELECT id, number, student_code, name, class_level, room, photo FROM students WHERE id=?',
            (sid,)).fetchone()
    if not s:
        return jsonify(error='ไม่พบนักเรียน'), 404
    return jsonify(dict(s))

@app.delete('/api/students/<int:sid>')
@admin_required
def api_student_delete(sid):
    with get_db() as con:
        con.execute('DELETE FROM students WHERE id=?', (sid,))
        audit_log(con, 'delete_student', 'student', sid, None)
    return jsonify(success=True)

@app.post('/api/students')
@login_required
def api_student_create():
    """เพิ่มนักเรียนทีละคน (admin หรือครูประจำชั้น)"""
    u = current_user()
    b = request.get_json() or {}
    name = (b.get('name') or '').strip()
    if not name:
        return jsonify(success=False, message='กรอกชื่อนักเรียน'), 400

    try:
        level = int(b.get('class_level'))
        room = int(b.get('room'))
    except (TypeError, ValueError):
        return jsonify(success=False, message='ชั้นและห้องไม่ถูกต้อง'), 400

    if not (1 <= level <= 6) or room < 1:
        return jsonify(success=False, message='ชั้นต้องเป็น 1-6, ห้องต้องเป็นเลขบวก'), 400

    # ครูสร้างได้เฉพาะนักเรียนในห้องตัวเอง
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if level != u['assigned_level'] or (u.get('assigned_room') and room != u['assigned_room']):
            return jsonify(success=False, message='ไม่มีสิทธิ์เพิ่มในห้องนี้'), 403

    number = b.get('number')
    student_code = (b.get('student_code') or '').strip() or None
    gender = (b.get('gender') or '').strip() or None
    birthdate = (b.get('birthdate') or '').strip() or None
    national_id = (b.get('national_id') or '').strip() or None
    try:
        number = int(number) if number else None
    except (TypeError, ValueError):
        number = None

    # Validate
    if national_id and (not national_id.isdigit() or len(national_id) != 13):
        return jsonify(success=False, message='เลขบัตรประชาชนต้องเป็นตัวเลข 13 หลัก'), 400
    if birthdate:
        try: datetime.date.fromisoformat(birthdate)
        except ValueError: return jsonify(success=False, message='วันเกิดไม่ถูกต้อง (YYYY-MM-DD)'), 400

    with get_db() as con:
        cur = con.execute(
            'INSERT INTO students (number, student_code, name, class_level, room, gender, birthdate, national_id) '
            'VALUES (?,?,?,?,?,?,?,?)',
            (number, student_code, name, level, room, gender, birthdate, national_id)
        )
        sid = cur.lastrowid
        audit_log(con, 'create_student', 'student', sid,
                  {'name': name, 'class': f'{level}/{room}'})

    return jsonify(success=True, id=sid, message=f'เพิ่มนักเรียน "{name}" สำเร็จ')

@app.put('/api/students/<int:sid>')
@login_required
def api_student_update(sid):
    """อัพเดทข้อมูลนักเรียน — admin หรือครูประจำชั้น"""
    u = current_user()
    b = request.get_json() or {}

    with get_db() as con:
        s = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not s:
            return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        # ครูแก้ได้เฉพาะห้องตัวเอง
        if u['role'] == 'teacher' and u.get('assigned_level'):
            if s['class_level'] != u['assigned_level']:
                return jsonify(success=False, message='ไม่มีสิทธิ์แก้ไขห้องนี้'), 403
            if u.get('assigned_room') and s['room'] != u['assigned_room']:
                return jsonify(success=False, message='ไม่มีสิทธิ์แก้ไขห้องนี้'), 403

        # อ่านค่าเดิมเป็น default — ถ้าไม่ส่งมาให้เก็บเดิม
        updates = {}
        if 'name' in b:
            name = (b.get('name') or '').strip()
            if not name: return jsonify(success=False, message='ชื่อห้ามว่าง'), 400
            updates['name'] = name
        if 'number' in b:
            try: updates['number'] = int(b['number']) if b['number'] else None
            except (TypeError, ValueError): pass
        if 'student_code' in b:
            updates['student_code'] = (b.get('student_code') or '').strip() or None
        if 'class_level' in b:
            try:
                lv = int(b['class_level'])
                if not (1 <= lv <= 6): raise ValueError()
                # ครูเปลี่ยน level ไม่ได้
                if u['role'] != 'admin' and lv != s['class_level']:
                    return jsonify(success=False, message='ครูเปลี่ยนชั้นไม่ได้'), 403
                updates['class_level'] = lv
            except (TypeError, ValueError):
                return jsonify(success=False, message='ชั้นไม่ถูกต้อง'), 400
        if 'room' in b:
            try:
                rm = int(b['room'])
                if rm < 1: raise ValueError()
                if u['role'] != 'admin' and rm != s['room']:
                    return jsonify(success=False, message='ครูเปลี่ยนห้องไม่ได้'), 403
                updates['room'] = rm
            except (TypeError, ValueError):
                return jsonify(success=False, message='ห้องไม่ถูกต้อง'), 400
        if 'gender' in b:
            updates['gender'] = (b.get('gender') or '').strip() or None
        if 'birthdate' in b:
            bd = (b.get('birthdate') or '').strip() or None
            if bd:
                try: datetime.date.fromisoformat(bd)
                except ValueError: return jsonify(success=False, message='วันเกิดไม่ถูกต้อง (YYYY-MM-DD)'), 400
            updates['birthdate'] = bd
        if 'national_id' in b:
            nid = (b.get('national_id') or '').strip() or None
            if nid and (not nid.isdigit() or len(nid) != 13):
                return jsonify(success=False, message='เลขบัตรประชาชนต้องเป็นตัวเลข 13 หลัก'), 400
            updates['national_id'] = nid
        if 'parent_name' in b:
            updates['parent_name'] = (b.get('parent_name') or '').strip() or None
        if 'parent_relation' in b:
            updates['parent_relation'] = (b.get('parent_relation') or '').strip() or None
        if 'parent_phone' in b:
            updates['parent_phone'] = (b.get('parent_phone') or '').strip() or None
        if 'address' in b:
            updates['address'] = (b.get('address') or '').strip() or None

        if not updates:
            return jsonify(success=False, message='ไม่มีข้อมูลให้อัพเดท'), 400

        cols = ', '.join(f'{k}=?' for k in updates.keys())
        con.execute(f'UPDATE students SET {cols} WHERE id=?', list(updates.values()) + [sid])
        audit_log(con, 'update_student', 'student', sid, {'fields': list(updates.keys())})

    return jsonify(success=True, message='อัพเดทข้อมูลสำเร็จ')

# ── ATTENDANCE ────────────────────────────────────────────

@app.get('/api/attendance')
@login_required
def api_attendance_get():
    u = current_user()
    level = request.args.get('level')
    room  = request.args.get('room')
    date  = request.args.get('date')
    if not all([level, room, date]):
        return jsonify(error='ต้องระบุ level, room และ date'), 400

    # Enforce assigned class for teachers
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if int(level) != u['assigned_level'] or (u.get('assigned_room') and int(room) != u['assigned_room']):
            return jsonify(error='ไม่มีสิทธิ์เข้าถึงห้องนี้'), 403

    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))

    with get_db() as con:
        rows = con.execute("""
            SELECT s.id, s.number, s.student_code, s.name, s.gender,
                   a.status AS status,
                   a.note,
                   CASE WHEN a.id IS NOT NULL THEN 1 ELSE 0 END AS is_recorded,
                   COALESCE((SELECT SUM(points) FROM behavior_logs WHERE student_id=s.id), 0) AS score_delta
            FROM students s
            LEFT JOIN attendance a ON s.id=a.student_id AND a.date=?
            WHERE s.class_level=? AND s.room=?
            ORDER BY s.number, s.name
        """, (date, int(level), int(room))).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        d['score'] = start_score + (d['score_delta'] or 0)
        result.append(d)
    return jsonify(result)

@app.post('/api/attendance')
@login_required
def api_attendance_post():
    body    = request.get_json()
    date    = body.get('date')
    records = body.get('records', [])
    if not date or not records:
        return jsonify(success=False, message='ข้อมูลไม่ถูกต้อง'), 400

    settings = get_settings()
    u = current_user()
    uid = u['id']

    # Verify all records belong to a class the user can access
    if u['role'] == 'teacher' and u.get('assigned_level'):
        ids = [r['student_id'] for r in records]
        ph = ','.join('?' * len(ids))
        with get_db() as con:
            owned = con.execute(
                f"""SELECT id FROM students WHERE id IN ({ph})
                    AND class_level=? {'AND room=?' if u.get('assigned_room') else ''}""",
                ids + [u['assigned_level']] + ([u['assigned_room']] if u.get('assigned_room') else [])
            ).fetchall()
        if len(owned) != len(ids):
            return jsonify(success=False, message='ไม่มีสิทธิ์บันทึกบางรายการ'), 403

    with get_db() as con:
        for r in records:
            cur = con.execute("""
                INSERT INTO attendance (student_id, date, status, note, recorded_by)
                VALUES (?,?,?,?,?)
                ON CONFLICT(student_id, date) DO UPDATE SET
                  status=excluded.status,
                  note=excluded.note,
                  recorded_by=excluded.recorded_by,
                  recorded_at=datetime('now','localtime')
            """, (r['student_id'], date, r['status'], r.get('note'), uid))
            att_id = con.execute('SELECT id FROM attendance WHERE student_id=? AND date=?',
                                 (r['student_id'], date)).fetchone()['id']
            apply_attendance_behavior(con, r['student_id'], att_id, date, r['status'], settings, uid)

    return jsonify(success=True, message=f"บันทึกการเช็คชื่อ {len(records)} คน สำเร็จ")

# ── UNIFORM CHECK (ตรวจเครื่องแต่งกาย) ────────────────────

# Default categories used by uniform check page. ใช้ behavior_rules ที่หมวด 'การแต่งกาย' เป็นหลัก
UNIFORM_CATEGORIES = ['การแต่งกาย']

@app.get('/api/uniform-check/rules')
@login_required
def api_uniform_rules():
    """รายการพฤติกรรมที่ใช้ในการตรวจเครื่องแต่งกาย (กรองจาก behavior_rules)"""
    cats_param = request.args.get('categories', ','.join(UNIFORM_CATEGORIES))
    cats = [c.strip() for c in cats_param.split(',') if c.strip()]
    if not cats:
        return jsonify([])
    placeholders = ','.join('?' * len(cats))
    with get_db() as con:
        rows = con.execute(
            f"""SELECT id, name, points, category FROM behavior_rules
                WHERE active=1 AND category IN ({placeholders})
                ORDER BY category, sort_order, id""",
            cats
        ).fetchall()
    return jsonify(rows_to_list(rows))

@app.get('/api/uniform-check/categories')
@login_required
def api_uniform_categories():
    """รายการหมวดทั้งหมดที่มีใน behavior_rules พร้อมจำนวน rules ในแต่ละหมวด"""
    with get_db() as con:
        rows = con.execute("""
            SELECT category, COUNT(*) AS count
            FROM behavior_rules
            WHERE active=1 AND category IS NOT NULL AND category != ''
            GROUP BY category
            ORDER BY MIN(sort_order), category
        """).fetchall()
    return jsonify(rows_to_list(rows))

@app.get('/api/uniform-check')
@login_required
def api_uniform_get():
    """ดึงรายชื่อนักเรียน + รายการที่ตรวจไปแล้วในวันนั้น"""
    u = current_user()
    level = request.args.get('level')
    room = request.args.get('room')
    date = request.args.get('date', today_iso())
    if not level or not room:
        return jsonify(error='ต้องระบุชั้น/ห้อง'), 400

    # Enforce assigned class for teachers
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if int(level) != u['assigned_level'] or (u.get('assigned_room') and int(room) != u['assigned_room']):
            return jsonify(error='ไม่มีสิทธิ์เข้าถึงห้องนี้'), 403

    with get_db() as con:
        students = con.execute("""
            SELECT id, number, student_code, name, gender
            FROM students WHERE class_level=? AND room=?
            ORDER BY number, name
        """, (int(level), int(room))).fetchall()

        student_ids = [s['id'] for s in students]
        marks = {}
        notes = {}
        att = {}
        if student_ids:
            ph = ','.join('?' * len(student_ids))
            # Uniform marks (รายการที่ตรวจ)
            rows = con.execute(
                f"""SELECT student_id, source_id, reason, points, note
                    FROM behavior_logs
                    WHERE source='uniform' AND date=? AND student_id IN ({ph})""",
                [date] + student_ids
            ).fetchall()
            for r in rows:
                marks.setdefault(r['student_id'], []).append(dict(r))
                if r['note']:
                    notes[r['student_id']] = r['note']  # use last non-empty note
            # Attendance สำหรับวันนั้น
            arows = con.execute(
                f"""SELECT student_id, status, note
                    FROM attendance
                    WHERE date=? AND student_id IN ({ph})""",
                [date] + student_ids
            ).fetchall()
            for r in arows:
                att[r['student_id']] = {'status': r['status'], 'note': r['note']}

    result = []
    for s in students:
        d = dict(s)
        d['marked_rule_ids'] = [m['source_id'] for m in marks.get(s['id'], [])]
        d['marked_total'] = sum(m['points'] for m in marks.get(s['id'], []))
        d['note'] = notes.get(s['id'], '')
        d['attendance'] = att.get(s['id'])  # None ถ้ายังไม่เช็คชื่อ
        result.append(d)

    return jsonify(students=result, date=date)

@app.post('/api/uniform-check')
@login_required
def api_uniform_save():
    """บันทึกผลการตรวจเครื่องแต่งกาย
    Body: {
      date: 'YYYY-MM-DD',
      records: [{student_id, rule_ids: [1, 2, ...]}, ...]
    }
    Idempotent: ลบ uniform records เดิมของวันนั้น/นักเรียนนั้น ก่อน insert ใหม่
    """
    u = current_user()
    b = request.get_json() or {}
    date = b.get('date')
    records = b.get('records', [])
    if not date or not isinstance(records, list):
        return jsonify(success=False, message='ข้อมูลไม่ถูกต้อง'), 400

    if not records:
        return jsonify(success=True, count=0, total_deduction=0, message='ไม่มีรายการ')

    # Verify access for teachers
    sid_list = [r.get('student_id') for r in records if r.get('student_id')]
    if not sid_list:
        return jsonify(success=False, message='ไม่มีนักเรียน'), 400

    with get_db() as con:
        ph = ','.join('?' * len(sid_list))
        students = {s['id']: dict(s) for s in con.execute(
            f'SELECT * FROM students WHERE id IN ({ph})', sid_list
        ).fetchall()}

        for sid in sid_list:
            s = students.get(sid)
            if not s or not can_access_student(u, s):
                return jsonify(success=False, message='ไม่มีสิทธิ์'), 403

        # Get all rule_ids referenced
        all_rule_ids = set()
        for rec in records:
            for rid in (rec.get('rule_ids') or []):
                all_rule_ids.add(int(rid))

        rules = {}
        if all_rule_ids:
            ph = ','.join('?' * len(all_rule_ids))
            for r in con.execute(
                f'SELECT id, name, points FROM behavior_rules WHERE id IN ({ph})',
                list(all_rule_ids)
            ).fetchall():
                rules[r['id']] = dict(r)

        # Clear existing uniform records for these students on this date
        con.execute(
            f"DELETE FROM behavior_logs WHERE source='uniform' AND date=? AND student_id IN ({','.join('?' * len(sid_list))})",
            [date] + sid_list
        )

        count = 0
        total_deduction = 0
        for rec in records:
            sid = rec['student_id']
            rule_ids = rec.get('rule_ids') or []
            note = (rec.get('note') or '').strip() or None
            for rid in rule_ids:
                rule = rules.get(int(rid))
                if not rule: continue
                con.execute(
                    """INSERT INTO behavior_logs (student_id, date, points, reason, source, source_id, note, recorded_by)
                       VALUES (?, ?, ?, ?, 'uniform', ?, ?, ?)""",
                    (sid, date, rule['points'], rule['name'], rule['id'], note, u['id'])
                )
                count += 1
                total_deduction += rule['points']

        audit_log(con, 'uniform_check', 'student', None,
                  {'date': date, 'students': len(sid_list), 'marks': count, 'total': total_deduction})

    return jsonify(success=True, count=count, total_deduction=total_deduction,
                   message=f'บันทึกการตรวจ {len(sid_list)} คน — รวมหัก {total_deduction} คะแนน')

# ── LATE STUDENTS + MAKEUP (บำเพ็ญประโยชน์แก้มาสาย) ──────

@app.post('/api/attendance/scan')
@login_required
def api_attendance_scan():
    """สแกน QR เพื่อบันทึก attendance ของนักเรียน 1 คน
    Body: { student_id: int, status: 'late' | 'present' | ..., date: 'YYYY-MM-DD' (optional, default=today) }
    Return: { success, student: {name, class, room, number, photo}, was_already }
    """
    u = current_user()
    b = request.get_json() or {}
    sid = b.get('student_id')
    status = b.get('status', 'late')
    date = b.get('date', today_iso())

    if not sid:
        return jsonify(success=False, message='ไม่พบ student_id'), 400
    if status not in ('present', 'absent', 'late', 'leave', 'activity'):
        return jsonify(success=False, message='สถานะไม่ถูกต้อง'), 400

    settings = get_settings()
    try: sid = int(sid)
    except (TypeError, ValueError):
        return jsonify(success=False, message='student_id ไม่ถูกต้อง'), 400

    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student:
            return jsonify(success=False, message='ไม่พบนักเรียน'), 404

        # ครูเวรเช็ค "มาสาย" ได้ทุกห้อง (งานหน้าประตู) — สถานะอื่นจำกัดเฉพาะห้องที่ปรึกษา
        if status != 'late' and u['role'] == 'teacher' and u.get('assigned_level'):
            if student['class_level'] != u['assigned_level']:
                return jsonify(success=False, message='ไม่มีสิทธิ์เช็คชื่อห้องนี้ (เช็คได้เฉพาะมาสาย)'), 403
            if u.get('assigned_room') and student['room'] != u['assigned_room']:
                return jsonify(success=False, message='ไม่มีสิทธิ์เช็คชื่อห้องนี้ (เช็คได้เฉพาะมาสาย)'), 403

        # ตรวจสอบว่ามี attendance ของวันนี้แล้วหรือยัง
        existing = con.execute(
            'SELECT id, status FROM attendance WHERE student_id=? AND date=?',
            (sid, date)
        ).fetchone()
        was_already = existing is not None
        previous_status = existing['status'] if existing else None

        # Insert หรือ update
        con.execute("""
            INSERT INTO attendance (student_id, date, status, recorded_by)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(student_id, date) DO UPDATE SET
              status=excluded.status,
              recorded_by=excluded.recorded_by,
              recorded_at=datetime('now','localtime')
        """, (sid, date, status, u['id']))

        att_id = con.execute('SELECT id FROM attendance WHERE student_id=? AND date=?',
                             (sid, date)).fetchone()['id']
        apply_attendance_behavior(con, sid, att_id, date, status, settings, u['id'])
        audit_log(con, 'scan_attendance', 'student', sid,
                  {'status': status, 'date': date, 'was_already': was_already})

    return jsonify(
        success=True,
        student={
            'id': student['id'],
            'name': student['name'],
            'number': student['number'],
            'student_code': student['student_code'],
            'class_level': student['class_level'],
            'room': student['room'],
            'photo': student['photo'],
        },
        date=date,
        status=status,
        was_already=was_already,
        previous_status=previous_status,
        message=f"บันทึก {student['name']} เป็น {status}"
    )

@app.delete('/api/attendance/scan')
@login_required
def api_attendance_scan_undo():
    """ลบรายการสแกนผิด — ลบ attendance + behavior_logs ที่เกี่ยวข้อง
    Body: { student_id: int, date: 'YYYY-MM-DD' (optional, default=today) }
    """
    u = current_user()
    b = request.get_json() or {}
    sid = b.get('student_id')
    date = b.get('date', today_iso())

    if not sid:
        return jsonify(success=False, message='ไม่พบ student_id'), 400
    try: sid = int(sid)
    except (TypeError, ValueError):
        return jsonify(success=False, message='student_id ไม่ถูกต้อง'), 400

    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student:
            return jsonify(success=False, message='ไม่พบนักเรียน'), 404

        # ครูที่มี assigned_room ทำได้เฉพาะห้องตัวเอง
        if u['role'] == 'teacher' and u.get('assigned_level'):
            if student['class_level'] != u['assigned_level']:
                return jsonify(success=False, message='ไม่มีสิทธิ์ลบรายการห้องนี้'), 403
            if u.get('assigned_room') and student['room'] != u['assigned_room']:
                return jsonify(success=False, message='ไม่มีสิทธิ์ลบรายการห้องนี้'), 403

        att = con.execute('SELECT id FROM attendance WHERE student_id=? AND date=?',
                          (sid, date)).fetchone()
        if not att:
            return jsonify(success=False, message='ไม่พบรายการเช็คชื่อ'), 404

        # ลบ behavior_logs ที่เกิดจาก attendance นี้ (จะคืนคะแนนอัตโนมัติ)
        con.execute('DELETE FROM behavior_logs WHERE source=? AND source_id=?',
                    ('attendance', att['id']))
        # ลบ attendance
        con.execute('DELETE FROM attendance WHERE id=?', (att['id'],))
        audit_log(con, 'undo_scan', 'student', sid, {'date': date})

    return jsonify(success=True, message=f"ลบรายการของ {student['name']} แล้ว")

@app.get('/api/late')
@login_required
def api_late_list():
    """รายชื่อนักเรียนที่มาสายในวันที่ระบุ + สถานะการบำเพ็ญประโยชน์"""
    u = current_user()
    date = request.args.get('date', today_iso())
    where, params, _, _ = user_class_filter(u)

    with get_db() as con:
        rows = con.execute(f"""
            SELECT a.id AS attendance_id, a.date, a.note,
                   s.id, s.number, s.student_code, s.name, s.class_level, s.room,
                   mk.id AS makeup_id, mk.points AS makeup_points, mk.created_at AS makeup_at,
                   u.full_name AS makeup_by_name
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            LEFT JOIN behavior_logs mk ON mk.source='makeup' AND mk.source_id=a.id
            LEFT JOIN users u ON u.id = mk.recorded_by
            WHERE a.status='late' AND a.date=? {where}
            ORDER BY s.class_level, s.room, s.number, s.name
        """, [date] + params).fetchall()
    return jsonify(rows=rows_to_list(rows), date=date)

@app.post('/api/late/<int:attendance_id>/makeup')
@login_required
def api_late_makeup(attendance_id):
    """บันทึกว่านักเรียนทำบำเพ็ญประโยชน์แล้ว → คืนคะแนนตามที่ตั้งไว้"""
    u = current_user()
    settings = get_settings()
    # ใช้ makeup_late_points (กำหนดได้ในหน้า settings) — fallback เป็น deduct_late ถ้าไม่ได้ตั้ง
    refund = int(settings.get('makeup_late_points') or settings.get('deduct_late', '1') or 0)

    with get_db() as con:
        att = con.execute("""
            SELECT a.*, s.class_level, s.room
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE a.id=? AND a.status='late'
        """, (attendance_id,)).fetchone()
        if not att:
            return jsonify(success=False, message='ไม่พบรายการมาสาย'), 404
        if not can_access_student(u, att):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403

        # ตรวจสอบว่ายังไม่ได้แก้
        existing = con.execute(
            "SELECT id FROM behavior_logs WHERE source='makeup' AND source_id=?",
            (attendance_id,)
        ).fetchone()
        if existing:
            return jsonify(success=False, message='นักเรียนคนนี้บำเพ็ญประโยชน์แล้ว'), 400

        # บันทึกการคืนคะแนน
        con.execute("""
            INSERT INTO behavior_logs (student_id, date, points, reason, source, source_id, recorded_by)
            VALUES (?, ?, ?, 'บำเพ็ญประโยชน์ (แก้มาสาย)', 'makeup', ?, ?)
        """, (att['student_id'], today_iso(), refund, attendance_id, u['id']))
        audit_log(con, 'late_makeup', 'attendance', attendance_id,
                  {'student_id': att['student_id'], 'refund': refund})

    return jsonify(success=True, refund=refund,
                   message=f'บันทึกการบำเพ็ญประโยชน์ — คืนคะแนน +{refund}')

@app.delete('/api/late/<int:attendance_id>/makeup')
@login_required
def api_late_undo_makeup(attendance_id):
    """ยกเลิกการบำเพ็ญประโยชน์ (เช่นบันทึกผิดคน)"""
    u = current_user()
    with get_db() as con:
        att = con.execute("""
            SELECT a.*, s.class_level, s.room
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE a.id=?
        """, (attendance_id,)).fetchone()
        if not att or not can_access_student(u, att):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        con.execute("DELETE FROM behavior_logs WHERE source='makeup' AND source_id=?", (attendance_id,))
        audit_log(con, 'late_undo_makeup', 'attendance', attendance_id, None)
    return jsonify(success=True, message='ยกเลิกการบำเพ็ญแล้ว')

@app.get('/api/line-summary')
@login_required
def api_line_summary():
    """ดึงข้อมูลสำหรับสร้างข้อความสรุปรายวันส่งใน LINE กลุ่มผู้ปกครอง"""
    u = current_user()
    level = request.args.get('level')
    room = request.args.get('room')
    date = request.args.get('date', today_iso())
    if not level or not room:
        return jsonify(error='ต้องระบุชั้น/ห้อง'), 400

    # Enforce assigned class for teachers
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if int(level) != u['assigned_level'] or (u.get('assigned_room') and int(room) != u['assigned_room']):
            return jsonify(error='ไม่มีสิทธิ์เข้าถึงห้องนี้'), 403

    settings = get_settings()
    with get_db() as con:
        # นักเรียนทั้งหมดในห้อง
        students = con.execute("""
            SELECT id, number, name FROM students
            WHERE class_level=? AND room=?
            ORDER BY number, name
        """, (int(level), int(room))).fetchall()

        # Attendance ของวันนั้น
        att_rows = con.execute("""
            SELECT a.student_id, a.status, a.note,
                   mk.id AS makeup_id
            FROM attendance a
            LEFT JOIN behavior_logs mk ON mk.source='makeup' AND mk.source_id=a.id
            WHERE a.date=? AND a.student_id IN (
                SELECT id FROM students WHERE class_level=? AND room=?
            )
        """, (date, int(level), int(room))).fetchall()
        att_map = {r['student_id']: dict(r) for r in att_rows}

        # Uniform check ของวันนั้น
        uniform_rows = con.execute("""
            SELECT student_id, reason, points
            FROM behavior_logs
            WHERE source='uniform' AND date=? AND student_id IN (
                SELECT id FROM students WHERE class_level=? AND room=?
            )
            ORDER BY student_id
        """, (date, int(level), int(room))).fetchall()
        uniform_map = {}
        for r in uniform_rows:
            uniform_map.setdefault(r['student_id'], []).append({'reason': r['reason'], 'points': r['points']})

    # Categorize students
    present, absent, late, leave, activity, unmarked = [], [], [], [], [], []
    for s in students:
        att = att_map.get(s['id'])
        if not att:
            unmarked.append(dict(s))
            continue
        info = dict(s, note=att.get('note'), made_up=bool(att.get('makeup_id')))
        if att['status'] == 'present': present.append(info)
        elif att['status'] == 'absent': absent.append(info)
        elif att['status'] == 'late': late.append(info)
        elif att['status'] == 'leave': leave.append(info)
        elif att['status'] == 'activity': activity.append(info)

    # Uniform violations
    uniform_violations = []
    for sid, items in uniform_map.items():
        s = next((dict(x) for x in students if x['id'] == sid), None)
        if s:
            s['violations'] = items
            s['total_deduction'] = sum(i['points'] for i in items)
            uniform_violations.append(s)

    parent_url = f"{request.host_url.rstrip('/')}/parent.html"

    return jsonify(
        school_name=settings.get('school_name', 'โรงเรียน'),
        date=date,
        level=int(level), room=int(room),
        total=len(students),
        present=present, absent=absent, late=late,
        leave=leave, activity=activity, unmarked=unmarked,
        uniform_violations=uniform_violations,
        parent_url=parent_url,
        attendance_checked=len(att_rows) > 0
    )

@app.get('/api/late/pending')
@login_required
def api_late_pending():
    """รายชื่อนักเรียนที่ค้างบำเพ็ญประโยชน์ (มาสายแต่ยังไม่ได้บำเพ็ญ) — ทุกวัน รวมวันนี้"""
    u = current_user()
    where, params, _, _ = user_class_filter(u)
    from_date = request.args.get('from')
    if not from_date:
        from_date = (datetime.date.today() - datetime.timedelta(days=60)).isoformat()

    with get_db() as con:
        rows = con.execute(f"""
            SELECT a.id AS attendance_id, a.date, a.note,
                   s.id, s.number, s.student_code, s.name, s.class_level, s.room
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            LEFT JOIN behavior_logs mk ON mk.source='makeup' AND mk.source_id=a.id
            WHERE a.status='late' AND mk.id IS NULL
              AND a.date >= ? AND a.date <= ?
              {where}
            ORDER BY a.date ASC, s.class_level, s.room, s.number, s.name
        """, [from_date, today_iso()] + params).fetchall()
    return jsonify(rows=rows_to_list(rows), from_date=from_date, until=today_iso())

@app.get('/api/late/summary')
@login_required
def api_late_summary():
    """สรุปจำนวนนักเรียนมาสาย + บำเพ็ญแล้ว/ยัง — สำหรับ dashboard"""
    u = current_user()
    date = request.args.get('date', today_iso())
    where, params, _, _ = user_class_filter(u)
    with get_db() as con:
        row = con.execute(f"""
            SELECT
              COUNT(*) AS total_late,
              SUM(CASE WHEN mk.id IS NOT NULL THEN 1 ELSE 0 END) AS made_up,
              SUM(CASE WHEN mk.id IS NULL THEN 1 ELSE 0 END) AS pending
            FROM attendance a
            JOIN students s ON s.id=a.student_id
            LEFT JOIN behavior_logs mk ON mk.source='makeup' AND mk.source_id=a.id
            WHERE a.status='late' AND a.date=? {where}
        """, [date] + params).fetchone()
    result = dict(row)
    result['date'] = date
    return jsonify(result)

# ── BEHAVIOR SCORES ───────────────────────────────────────

@app.get('/api/behavior/<int:sid>')
@login_required
def api_behavior_get(sid):
    u = current_user()
    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student:
            return jsonify(error='not found'), 404
        if not can_access_student(u, student):
            return jsonify(error='forbidden'), 403
        logs = con.execute("""
            SELECT bl.*, u.full_name as recorded_by_name
            FROM behavior_logs bl
            LEFT JOIN users u ON bl.recorded_by=u.id
            WHERE bl.student_id=? ORDER BY bl.date DESC, bl.id DESC
        """, (sid,)).fetchall()
        delta = con.execute('SELECT COALESCE(SUM(points),0) AS s FROM behavior_logs WHERE student_id=?',
                            (sid,)).fetchone()['s']
    return jsonify(
        student=dict(student),
        score=start_score + delta,
        start_score=start_score,
        logs=rows_to_list(logs)
    )

@app.post('/api/behavior')
@login_required
def api_behavior_add():
    u = current_user()
    b = request.get_json() or {}
    sid    = b.get('student_id')
    points = b.get('points')
    reason = (b.get('reason') or '').strip()
    if not sid or points is None or not reason:
        return jsonify(success=False, message='กรอกข้อมูลให้ครบ'), 400
    try:
        points = int(points)
    except ValueError:
        return jsonify(success=False, message='คะแนนต้องเป็นตัวเลข'), 400

    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student:
            return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์เข้าถึงนักเรียนคนนี้'), 403
        con.execute("""
            INSERT INTO behavior_logs (student_id, date, points, reason, source, recorded_by)
            VALUES (?, ?, ?, ?, 'manual', ?)
        """, (sid, today_iso(), points, reason, u['id']))
    return jsonify(success=True, message='บันทึกคะแนนแล้ว')

@app.delete('/api/behavior/<int:log_id>')
@login_required
def api_behavior_delete(log_id):
    with get_db() as con:
        log = con.execute('SELECT source FROM behavior_logs WHERE id=?', (log_id,)).fetchone()
        if not log:
            return jsonify(success=False, message='ไม่พบรายการ'), 404
        if log['source'] != 'manual':
            return jsonify(success=False, message='ลบเฉพาะรายการที่บันทึกเองได้'), 400
        con.execute('DELETE FROM behavior_logs WHERE id=?', (log_id,))
    return jsonify(success=True)

# ── BEHAVIOR RULES (พฤติกรรมผิดระเบียบ) ───────────────────

@app.get('/api/behavior-rules')
@login_required
def api_rules_list():
    only_active = request.args.get('active') == '1'
    sql = 'SELECT * FROM behavior_rules'
    if only_active:
        sql += ' WHERE active = 1'
    sql += ' ORDER BY sort_order, id'
    with get_db() as con:
        rows = con.execute(sql).fetchall()
    return jsonify(rows_to_list(rows))

@app.post('/api/behavior-rules')
@admin_required
def api_rules_create():
    b = request.get_json() or {}
    name     = (b.get('name') or '').strip()
    category = (b.get('category') or '').strip() or None
    sort_order = int(b.get('sort_order') or 999)
    try:
        points = int(b.get('points'))
    except (TypeError, ValueError):
        return jsonify(success=False, message='คะแนนต้องเป็นตัวเลข'), 400
    if not name:
        return jsonify(success=False, message='กรอกชื่อพฤติกรรม'), 400

    with get_db() as con:
        cur = con.execute(
            'INSERT INTO behavior_rules (name, points, category, sort_order) VALUES (?,?,?,?)',
            (name, points, category, sort_order)
        )
    return jsonify(success=True, id=cur.lastrowid, message='เพิ่มพฤติกรรมสำเร็จ')

@app.put('/api/behavior-rules/<int:rid>')
@admin_required
def api_rules_update(rid):
    b = request.get_json() or {}
    fields, params = [], []
    for k in ('name', 'category'):
        if k in b: fields.append(f'{k}=?'); params.append((b[k] or '').strip() or None)
    if 'points' in b:
        try: pts = int(b['points'])
        except: return jsonify(success=False, message='คะแนนต้องเป็นตัวเลข'), 400
        fields.append('points=?'); params.append(pts)
    if 'active' in b:
        fields.append('active=?'); params.append(1 if b['active'] else 0)
    if 'sort_order' in b:
        fields.append('sort_order=?'); params.append(int(b['sort_order']))
    if not fields:
        return jsonify(success=False, message='ไม่มีข้อมูลให้อัปเดต'), 400
    params.append(rid)
    with get_db() as con:
        con.execute(f'UPDATE behavior_rules SET {", ".join(fields)} WHERE id=?', params)
    return jsonify(success=True, message='แก้ไขสำเร็จ')

@app.delete('/api/behavior-rules/<int:rid>')
@admin_required
def api_rules_delete(rid):
    with get_db() as con:
        con.execute('DELETE FROM behavior_rules WHERE id=?', (rid,))
    return jsonify(success=True)

def build_behavior_filter(args, user=None):
    """Build filters; for teachers, forces their assigned class."""
    sem  = args.get('semester'); year = args.get('year')
    if sem and year:
        settings = get_settings()
        from_d, to_d = semester_range(int(year), int(sem), settings)
    else:
        from_d, to_d = args.get('from'), args.get('to')
    level  = args.get('level')
    room   = args.get('room')
    source = args.get('source')
    if user and user['role'] == 'teacher' and user.get('assigned_level'):
        level = user['assigned_level']
        if user.get('assigned_room'):
            room = user['assigned_room']
    return from_d, to_d, level, room, source

@app.get('/api/behavior/report')
@login_required
def api_behavior_report():
    """Summary: คะแนนคงเหลือของแต่ละคน + การหักในช่วงเวลา"""
    from_d, to_d, level, room, _ = build_behavior_filter(request.args, current_user())
    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))

    # Subquery: deductions in date range
    join_cond, jp = 's.id = bl.student_id', []
    if from_d: join_cond += ' AND bl.date >= ?'; jp.append(from_d)
    if to_d:   join_cond += ' AND bl.date <= ?'; jp.append(to_d)

    where_cond, wp = '1=1', []
    if level: where_cond += ' AND s.class_level=?'; wp.append(int(level))
    if room:  where_cond += ' AND s.room=?';        wp.append(int(room))

    sql = f"""
        SELECT s.id, s.number, s.student_code, s.name, s.class_level, s.room,
               COUNT(bl.id) AS event_count,
               COALESCE(SUM(bl.points), 0) AS range_delta,
               COALESCE(SUM(CASE WHEN bl.source='attendance' THEN bl.points ELSE 0 END), 0) AS auto_delta,
               COALESCE(SUM(CASE WHEN bl.source='manual'     THEN bl.points ELSE 0 END), 0) AS manual_delta
        FROM students s
        LEFT JOIN behavior_logs bl ON {join_cond}
        WHERE {where_cond}
        GROUP BY s.id
        ORDER BY s.class_level, s.room, s.number, s.name
    """
    with get_db() as con:
        rows = con.execute(sql, jp + wp).fetchall()
        # total score (all-time, not just range)
        totals = {r['student_id']: r['total'] for r in con.execute(
            'SELECT student_id, SUM(points) AS total FROM behavior_logs GROUP BY student_id'
        ).fetchall()}

    out = []
    for r in rows:
        d = dict(r)
        d['current_score'] = start_score + (totals.get(d['id'], 0) or 0)
        out.append(d)
    return jsonify(rows=out, range={'from': from_d, 'to': to_d}, start_score=start_score)

@app.get('/api/behavior/logs')
@login_required
def api_behavior_logs():
    """ประวัติย้อนหลัง: รายการหักคะแนนทุกครั้ง"""
    from_d, to_d, level, room, source = build_behavior_filter(request.args, current_user())

    cond, params = '1=1', []
    if from_d: cond += ' AND bl.date >= ?'; params.append(from_d)
    if to_d:   cond += ' AND bl.date <= ?'; params.append(to_d)
    if level:  cond += ' AND s.class_level=?'; params.append(int(level))
    if room:   cond += ' AND s.room=?';        params.append(int(room))
    if source: cond += ' AND bl.source=?';     params.append(source)

    sql = f"""
        SELECT bl.id, bl.date, bl.points, bl.reason, bl.source, bl.created_at,
               s.id AS student_id, s.number, s.name, s.class_level, s.room,
               u.full_name AS recorded_by_name
        FROM behavior_logs bl
        JOIN students s ON bl.student_id = s.id
        LEFT JOIN users u ON bl.recorded_by = u.id
        WHERE {cond}
        ORDER BY bl.date DESC, bl.id DESC
    """
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()
    return jsonify(rows=rows_to_list(rows), range={'from': from_d, 'to': to_d})

@app.get('/api/behavior/report/export')
@login_required
def api_behavior_report_export():
    from_d, to_d, level, room, _ = build_behavior_filter(request.args, current_user())
    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))

    join_cond, jp = 's.id = bl.student_id', []
    if from_d: join_cond += ' AND bl.date >= ?'; jp.append(from_d)
    if to_d:   join_cond += ' AND bl.date <= ?'; jp.append(to_d)
    where_cond, wp = '1=1', []
    if level: where_cond += ' AND s.class_level=?'; wp.append(int(level))
    if room:  where_cond += ' AND s.room=?';        wp.append(int(room))

    sql = f"""
        SELECT s.id, s.number, s.student_code, s.name, s.class_level, s.room,
               COUNT(bl.id) AS event_count,
               COALESCE(SUM(bl.points), 0) AS range_delta,
               COALESCE(SUM(CASE WHEN bl.source='attendance' THEN bl.points ELSE 0 END), 0) AS auto_delta,
               COALESCE(SUM(CASE WHEN bl.source='manual'     THEN bl.points ELSE 0 END), 0) AS manual_delta
        FROM students s LEFT JOIN behavior_logs bl ON {join_cond}
        WHERE {where_cond} GROUP BY s.id
        ORDER BY s.class_level, s.room, s.number, s.name
    """
    with get_db() as con:
        rows = con.execute(sql, jp + wp).fetchall()
        totals = {r['student_id']: r['total'] for r in con.execute(
            'SELECT student_id, SUM(points) AS total FROM behavior_logs GROUP BY student_id'
        ).fetchall()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'คะแนนความประพฤติ'

    range_txt = ''
    if from_d and to_d: range_txt = f'ช่วงข้อมูล: {from_d} ถึง {to_d}'
    elif from_d:        range_txt = f'ตั้งแต่ {from_d}'
    elif to_d:          range_txt = f'ถึง {to_d}'

    headers = ['เลขที่','รหัสนักเรียน','ชื่อ-นามสกุล','ชั้น',
               'จำนวนครั้งที่หัก','คะแนนหักช่วงนี้','ออโต้','บันทึกเอง','คะแนนเริ่มต้น','คะแนนคงเหลือ']
    hfill = PatternFill(fill_type='solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        current = start_score + (totals.get(r['id'], 0) or 0)
        ws.append([
            r['number'], r['student_code'], r['name'],
            f"ม.{r['class_level']}/{r['room']}",
            r['event_count'], r['range_delta'], r['auto_delta'], r['manual_delta'],
            start_score, current
        ])

    widths = [8, 14, 30, 10, 14, 14, 10, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Add title row at the top (insert before row 1)
    ws.insert_rows(1)
    if range_txt:
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value='รายงานคะแนนความประพฤติ — โรงเรียนตะกั่วทุ่งงานทวีวิทยาคม').font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=range_txt)
    else:
        ws.cell(row=1, column=1, value='รายงานคะแนนความประพฤติ — โรงเรียนตะกั่วทุ่งงานทวีวิทยาคม').font = Font(bold=True, size=14)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"behavior_report_{today_iso()}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/api/behavior/logs/export')
@login_required
def api_behavior_logs_export():
    from_d, to_d, level, room, source = build_behavior_filter(request.args, current_user())
    cond, params = '1=1', []
    if from_d: cond += ' AND bl.date >= ?'; params.append(from_d)
    if to_d:   cond += ' AND bl.date <= ?'; params.append(to_d)
    if level:  cond += ' AND s.class_level=?'; params.append(int(level))
    if room:   cond += ' AND s.room=?';        params.append(int(room))
    if source: cond += ' AND bl.source=?';     params.append(source)

    sql = f"""
        SELECT bl.date, s.class_level, s.room, s.number, s.name, s.student_code,
               bl.points, bl.reason, bl.source, u.full_name AS recorded_by_name, bl.created_at
        FROM behavior_logs bl JOIN students s ON bl.student_id=s.id
        LEFT JOIN users u ON bl.recorded_by = u.id
        WHERE {cond}
        ORDER BY bl.date DESC, s.class_level, s.room, s.number
    """
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = 'ประวัติการหักคะแนน'

    headers = ['วันที่','ชั้น','เลขที่','รหัสนักเรียน','ชื่อ-นามสกุล',
               'คะแนน','เหตุผล','ประเภท','ผู้บันทึก','บันทึกเมื่อ']
    hfill = PatternFill(fill_type='solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            r['date'], f"ม.{r['class_level']}/{r['room']}", r['number'],
            r['student_code'], r['name'], r['points'], r['reason'],
            'อัตโนมัติ' if r['source']=='attendance' else 'บันทึกเอง',
            r['recorded_by_name'] or '-', r['created_at']
        ])

    for col, w in zip(ws.columns, [12, 10, 8, 14, 28, 10, 30, 12, 18, 18]):
        ws.column_dimensions[col[0].column_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"behavior_logs_{today_iso()}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/api/behavior/ranking')
@login_required
def api_behavior_ranking():
    u = current_user()
    settings = get_settings()
    start = int(settings.get('start_score', '100'))
    where, params, _, _ = user_class_filter(u)
    with get_db() as con:
        rows = con.execute(f"""
            SELECT s.id, s.number, s.name, s.class_level, s.room,
                   {start} + COALESCE(SUM(bl.points), 0) AS score,
                   COALESCE(SUM(bl.points), 0) AS delta
            FROM students s
            LEFT JOIN behavior_logs bl ON s.id=bl.student_id
            WHERE 1=1 {where}
            GROUP BY s.id
            HAVING delta < 0
            ORDER BY score ASC
            LIMIT 30
        """, params).fetchall()
    return jsonify(rows_to_list(rows))

# ── DASHBOARD ─────────────────────────────────────────────

@app.get('/api/dashboard')
@login_required
def api_dashboard():
    u = current_user()
    today = today_iso()
    settings = get_settings()
    threshold = int(settings.get('alert_threshold', '5'))

    today_d = datetime.date.today()
    month_start = today_d.replace(day=1).isoformat()

    where, params, _, _ = user_class_filter(u)

    with get_db() as con:
        total = con.execute(f'SELECT COUNT(*) AS n FROM students s WHERE 1=1 {where}',
                            params).fetchone()['n']

        stats = con.execute(f"""
            SELECT COUNT(*) AS checked,
                   SUM(status='present')  AS present,
                   SUM(status='absent')   AS absent,
                   SUM(status='late')     AS late,
                   SUM(status='leave')    AS leave,
                   SUM(status='activity') AS activity
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE a.date=? {where}
        """, [today] + params).fetchone()

        classes = con.execute(f"""
            SELECT s.class_level, s.room,
                   COUNT(DISTINCT s.id) AS total,
                   COUNT(DISTINCT a.student_id) AS checked
            FROM students s
            LEFT JOIN attendance a ON s.id=a.student_id AND a.date=?
            WHERE 1=1 {where}
            GROUP BY s.class_level, s.room
            ORDER BY s.class_level, s.room
        """, [today] + params).fetchall()

        alerts = con.execute(f"""
            SELECT s.id, s.number, s.name, s.class_level, s.room,
                   COUNT(*) AS absent_count
            FROM students s
            JOIN attendance a ON s.id=a.student_id
            WHERE a.status='absent' AND a.date >= ? {where}
            GROUP BY s.id
            HAVING absent_count >= ?
            ORDER BY absent_count DESC
            LIMIT 30
        """, [month_start] + params + [threshold]).fetchall()

        month_stats = con.execute(f"""
            SELECT
              SUM(status='present')  AS present,
              SUM(status='absent')   AS absent,
              SUM(status='late')     AS late,
              SUM(status='leave')    AS leave,
              SUM(status='activity') AS activity,
              COUNT(DISTINCT date)   AS days
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE a.date >= ? {where}
        """, [month_start] + params).fetchone()

        # ─── BEHAVIOR SCORE STATS ───
        start_score = int(settings.get('start_score', '100'))

        # Overall score stats (cumulative)
        score_stats = con.execute(f"""
            SELECT
              COUNT(s.id) AS total_students,
              SUM(CASE WHEN bl_sum.delta < 0 THEN 1 ELSE 0 END) AS deducted_count,
              COALESCE(AVG({start_score} + COALESCE(bl_sum.delta, 0)), {start_score}) AS avg_score,
              COALESCE(MIN({start_score} + COALESCE(bl_sum.delta, 0)), {start_score}) AS min_score
            FROM students s
            LEFT JOIN (
              SELECT student_id, SUM(points) AS delta
              FROM behavior_logs GROUP BY student_id
            ) bl_sum ON bl_sum.student_id = s.id
            WHERE 1=1 {where}
        """, params).fetchone()

        # Month deductions (manual + auto)
        month_bhv = con.execute(f"""
            SELECT
              COUNT(*) AS events,
              COALESCE(SUM(CASE WHEN bl.points < 0 THEN bl.points ELSE 0 END), 0) AS total_deduction,
              COALESCE(SUM(CASE WHEN bl.source='manual'     THEN 1 ELSE 0 END), 0) AS manual_events,
              COALESCE(SUM(CASE WHEN bl.source='attendance' THEN 1 ELSE 0 END), 0) AS auto_events
            FROM behavior_logs bl
            JOIN students s ON s.id = bl.student_id
            WHERE bl.date >= ? {where}
        """, [month_start] + params).fetchone()

        # Lowest-score students (ต้องดูแลเป็นพิเศษ)
        score_alert_threshold = int(start_score * 0.8)  # < 80% triggers alert
        low_scores = con.execute(f"""
            SELECT s.id, s.number, s.name, s.class_level, s.room,
                   {start_score} + COALESCE(bl_sum.delta, 0) AS score,
                   COALESCE(bl_sum.delta, 0) AS delta
            FROM students s
            JOIN (
              SELECT student_id, SUM(points) AS delta
              FROM behavior_logs GROUP BY student_id
            ) bl_sum ON bl_sum.student_id = s.id
            WHERE bl_sum.delta < 0 {where}
            ORDER BY score ASC LIMIT 10
        """, params).fetchall()

    return jsonify(
        total=total,
        today=dict(stats),
        classSummary=rows_to_list(classes),
        alerts=rows_to_list(alerts),
        alertThreshold=threshold,
        monthStats=dict(month_stats),
        monthStart=month_start,
        date=today,
        scopedUser={
            'role': u['role'],
            'assigned_level': u.get('assigned_level'),
            'assigned_room':  u.get('assigned_room'),
        },
        behavior={
            'startScore':    start_score,
            'avgScore':      round(score_stats['avg_score'] or start_score, 1),
            'minScore':      score_stats['min_score'] or start_score,
            'deductedCount': score_stats['deducted_count'] or 0,
            'totalStudents': score_stats['total_students'] or 0,
            'monthEvents':   month_bhv['events'] or 0,
            'monthDeduction': month_bhv['total_deduction'] or 0,
            'monthManual':   month_bhv['manual_events'] or 0,
            'monthAuto':     month_bhv['auto_events'] or 0,
            'lowScores':     rows_to_list(low_scores),
            'alertThreshold': score_alert_threshold,
        }
    )

# ── ปค. FORMS (save / load / edit) ────────────────────────

@app.get('/api/pc-forms')
@login_required
def api_pcforms_list():
    """รายการฟอร์ม ปค. ที่บันทึกไว้ — กรองตาม ?student_id= หรือ ?type="""
    u = current_user()
    student_id = request.args.get('student_id')
    form_type = request.args.get('type')
    where, params = '', []
    if student_id:
        where += ' AND f.student_id=?'; params.append(int(student_id))
    if form_type:
        where += ' AND f.form_type=?'; params.append(form_type)
    # ครูเห็นเฉพาะห้องตัวเอง
    if u['role'] == 'teacher' and u.get('assigned_level'):
        where += ' AND s.class_level=?'; params.append(u['assigned_level'])
        if u.get('assigned_room'):
            where += ' AND s.room=?'; params.append(u['assigned_room'])
    with get_db() as con:
        rows = con.execute(f"""
            SELECT f.id, f.form_type, f.student_id, f.title, f.created_at, f.updated_at,
                   s.name AS student_name, s.class_level, s.room, s.number,
                   u.full_name AS recorded_by_name
            FROM pc_forms f
            JOIN students s ON s.id=f.student_id
            LEFT JOIN users u ON u.id=f.recorded_by
            WHERE 1=1 {where}
            ORDER BY f.updated_at DESC
        """, params).fetchall()
    return jsonify(rows_to_list(rows))

@app.get('/api/pc-forms/<int:fid>')
@login_required
def api_pcform_get(fid):
    u = current_user()
    with get_db() as con:
        f = con.execute("""
            SELECT f.*, s.class_level, s.room FROM pc_forms f
            JOIN students s ON s.id=f.student_id WHERE f.id=?
        """, (fid,)).fetchone()
    if not f:
        return jsonify(error='ไม่พบฟอร์ม'), 404
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if f['class_level'] != u['assigned_level'] or (u.get('assigned_room') and f['room'] != u['assigned_room']):
            return jsonify(error='ไม่มีสิทธิ์'), 403
    d = dict(f)
    try: d['payload'] = json.loads(d['payload'] or '{}')
    except Exception: d['payload'] = {}
    return jsonify(d)

@app.post('/api/pc-forms')
@login_required
def api_pcform_save():
    """บันทึกฟอร์มใหม่ หรืออัพเดท (ถ้ามี id)"""
    u = current_user()
    b = request.get_json() or {}
    fid        = b.get('id')
    form_type  = (b.get('form_type') or '').strip()
    student_id = b.get('student_id')
    title      = (b.get('title') or '').strip() or None
    payload    = b.get('payload') or {}

    if form_type not in ('pc1', 'pc3', 'pc4', 'pc5'):
        return jsonify(success=False, message='ชนิดฟอร์มไม่ถูกต้อง'), 400
    if not student_id:
        return jsonify(success=False, message='ไม่พบนักเรียน'), 400

    payload_json = json.dumps(payload, ensure_ascii=False)
    with get_db() as con:
        student = con.execute('SELECT class_level, room FROM students WHERE id=?', (student_id,)).fetchone()
        if not student:
            return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if u['role'] == 'teacher' and u.get('assigned_level'):
            if student['class_level'] != u['assigned_level'] or (u.get('assigned_room') and student['room'] != u['assigned_room']):
                return jsonify(success=False, message='ไม่มีสิทธิ์ห้องนี้'), 403

        if fid:
            con.execute("""
                UPDATE pc_forms SET title=?, payload=?, updated_at=datetime('now','localtime')
                WHERE id=?
            """, (title, payload_json, fid))
            audit_log(con, 'update_pcform', 'pc_form', fid, {'type': form_type})
            saved_id = fid
        else:
            cur = con.execute("""
                INSERT INTO pc_forms (form_type, student_id, title, payload, recorded_by)
                VALUES (?,?,?,?,?)
            """, (form_type, student_id, title, payload_json, u['id']))
            saved_id = cur.lastrowid
            audit_log(con, 'create_pcform', 'pc_form', saved_id, {'type': form_type})

    return jsonify(success=True, id=saved_id, message='บันทึกฟอร์มเรียบร้อย')

@app.delete('/api/pc-forms/<int:fid>')
@login_required
def api_pcform_delete(fid):
    u = current_user()
    with get_db() as con:
        f = con.execute("""
            SELECT f.id, s.class_level, s.room FROM pc_forms f
            JOIN students s ON s.id=f.student_id WHERE f.id=?
        """, (fid,)).fetchone()
        if not f:
            return jsonify(success=False, message='ไม่พบฟอร์ม'), 404
        if u['role'] == 'teacher' and u.get('assigned_level'):
            if f['class_level'] != u['assigned_level'] or (u.get('assigned_room') and f['room'] != u['assigned_room']):
                return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        con.execute('DELETE FROM pc_forms WHERE id=?', (fid,))
        audit_log(con, 'delete_pcform', 'pc_form', fid, None)
    return jsonify(success=True, message='ลบฟอร์มแล้ว')

# ── HEATMAP (calendar grid) ───────────────────────────────

@app.get('/api/heatmap')
@login_required
def api_heatmap():
    """ข้อมูลการมาเรียนรายวัน 1 ปี สำหรับ heatmap calendar
    Query: ?days=365 (default), ?level=1, ?room=1, ?student_id=123
    Return: { from, to, days: [{date, total, present, late, absent, leave, activity, attended_pct}] }
    """
    u = current_user()
    days = int(request.args.get('days', '365'))
    days = max(30, min(days, 730))  # 1 month - 2 years
    student_id = request.args.get('student_id')
    level = request.args.get('level')
    room = request.args.get('room')

    today_d = datetime.date.today()
    from_d = today_d - datetime.timedelta(days=days - 1)

    # Build filter
    where = ''
    params = [from_d.isoformat()]
    if student_id:
        try:
            sid = int(student_id)
            where += ' AND a.student_id=?'
            params.append(sid)
        except ValueError: pass
    if level:
        try:
            where += ' AND s.class_level=?'
            params.append(int(level))
        except ValueError: pass
    if room:
        try:
            where += ' AND s.room=?'
            params.append(int(room))
        except ValueError: pass

    # ครูดูได้เฉพาะห้องตัวเอง
    if u['role'] == 'teacher' and u.get('assigned_level'):
        where += ' AND s.class_level=?'
        params.append(u['assigned_level'])
        if u.get('assigned_room'):
            where += ' AND s.room=?'
            params.append(u['assigned_room'])

    with get_db() as con:
        rows = con.execute(f"""
            SELECT a.date,
                   COUNT(*) AS total,
                   SUM(a.status='present')  AS present,
                   SUM(a.status='late')     AS late,
                   SUM(a.status='absent')   AS absent,
                   SUM(a.status='leave')    AS leave,
                   SUM(a.status='activity') AS activity
            FROM attendance a
            JOIN students s ON s.id=a.student_id
            WHERE a.date >= ? {where}
            GROUP BY a.date
            ORDER BY a.date
        """, params).fetchall()

    days_list = []
    for r in rows:
        total = r['total'] or 0
        attended = (r['present'] or 0) + (r['late'] or 0)
        pct = round((attended / total * 100), 1) if total > 0 else 0
        days_list.append({
            'date':     r['date'],
            'total':    total,
            'present':  r['present'] or 0,
            'late':     r['late']    or 0,
            'absent':   r['absent']  or 0,
            'leave':    r['leave']   or 0,
            'activity': r['activity']or 0,
            'attended_pct': pct,
        })

    return jsonify(
        from_date=from_d.isoformat(),
        to_date=today_d.isoformat(),
        days=days_list,
    )

# ── DIRECTOR VIEW (public, key-protected, read-only) ──────

@app.get('/api/director')
def api_director():
    """ภาพรวมโรงเรียนสำหรับ ผอ. — ไม่ต้อง login แต่ต้องมี key ที่ถูก
    รับ ?date=YYYY-MM-DD (optional, default=วันนี้) เพื่อดูย้อนหลัง
    """
    key = request.args.get('key', '').strip()
    settings = get_settings()
    expected = (settings.get('director_view_key') or '').strip()
    if not expected or key != expected:
        return jsonify(error='invalid key'), 403

    # รับ date จาก query, ถ้าไม่มี/ผิด → ใช้วันนี้
    today = request.args.get('date', '').strip() or today_iso()
    try:
        today_d = datetime.date.fromisoformat(today)
    except ValueError:
        today_d = datetime.date.today()
        today = today_d.isoformat()

    month_start = today_d.replace(day=1).isoformat()
    start_score = int(settings.get('start_score', '100'))

    # 14 วันย้อนหลังจากวันที่เลือก
    fourteen = (today_d - datetime.timedelta(days=13)).isoformat()

    with get_db() as con:
        total = con.execute('SELECT COUNT(*) AS n FROM students').fetchone()['n']

        today_stats = con.execute("""
            SELECT
              COUNT(*) AS checked,
              SUM(status='present')  AS present,
              SUM(status='absent')   AS absent,
              SUM(status='late')     AS late,
              SUM(status='leave')    AS leave,
              SUM(status='activity') AS activity
            FROM attendance WHERE date=?
        """, (today,)).fetchone()

        month_stats = con.execute("""
            SELECT
              SUM(status='present')  AS present,
              SUM(status='absent')   AS absent,
              SUM(status='late')     AS late,
              SUM(status='leave')    AS leave,
              SUM(status='activity') AS activity,
              COUNT(DISTINCT date)   AS days
            FROM attendance WHERE date >= ?
        """, (month_start,)).fetchone()

        # ภาพรวมแนวโน้ม 14 วัน
        trend = con.execute("""
            SELECT date,
                   SUM(status='present')  AS present,
                   SUM(status='absent')   AS absent,
                   SUM(status='late')     AS late,
                   SUM(status='leave')    AS leave,
                   SUM(status='activity') AS activity
            FROM attendance
            WHERE date >= ?
            GROUP BY date ORDER BY date
        """, (fourteen,)).fetchall()

        # สรุปรายชั้น (ม.1-ม.6) สำหรับวันนี้
        by_level = con.execute("""
            SELECT s.class_level,
                   COUNT(DISTINCT s.id) AS total,
                   COUNT(DISTINCT CASE WHEN a.status='present'  THEN s.id END) AS present,
                   COUNT(DISTINCT CASE WHEN a.status='absent'   THEN s.id END) AS absent,
                   COUNT(DISTINCT CASE WHEN a.status='late'     THEN s.id END) AS late,
                   COUNT(DISTINCT CASE WHEN a.status='leave'    THEN s.id END) AS leave,
                   COUNT(DISTINCT a.student_id) AS checked
            FROM students s
            LEFT JOIN attendance a ON s.id=a.student_id AND a.date=?
            GROUP BY s.class_level
            ORDER BY s.class_level
        """, (today,)).fetchall()

        # คะแนนความประพฤติเฉลี่ย
        score_stats = con.execute(f"""
            SELECT
              COUNT(s.id) AS total_students,
              SUM(CASE WHEN bl_sum.delta < 0 THEN 1 ELSE 0 END) AS deducted,
              COALESCE(AVG({start_score} + COALESCE(bl_sum.delta, 0)), {start_score}) AS avg_score,
              COALESCE(MIN({start_score} + COALESCE(bl_sum.delta, 0)), {start_score}) AS min_score
            FROM students s
            LEFT JOIN (
              SELECT student_id, SUM(points) AS delta
              FROM behavior_logs GROUP BY student_id
            ) bl_sum ON bl_sum.student_id = s.id
        """).fetchone()

        # นักเรียนคะแนนต่ำที่สุด 5 คน
        low_scores = con.execute(f"""
            SELECT s.number, s.name, s.class_level, s.room,
                   {start_score} + COALESCE(bl_sum.delta, 0) AS score
            FROM students s
            JOIN (
              SELECT student_id, SUM(points) AS delta
              FROM behavior_logs GROUP BY student_id
            ) bl_sum ON bl_sum.student_id = s.id
            WHERE bl_sum.delta < 0
            ORDER BY score ASC LIMIT 5
        """).fetchall()

        # นักเรียนขาดบ่อย (เกินกำหนด) เดือนนี้
        alert_threshold = int(settings.get('alert_threshold', '5'))
        frequent_absent = con.execute("""
            SELECT s.number, s.name, s.class_level, s.room,
                   COUNT(*) AS absent_count
            FROM students s
            JOIN attendance a ON s.id=a.student_id
            WHERE a.status='absent' AND a.date >= ?
            GROUP BY s.id
            HAVING absent_count >= ?
            ORDER BY absent_count DESC LIMIT 10
        """, (month_start, alert_threshold)).fetchall()

    return jsonify(
        schoolName=settings.get('school_name', ''),
        schoolLogo=settings.get('school_logo', ''),
        themeColor=settings.get('theme_color', '#1a5276'),
        directorName=settings.get('director_name', ''),
        directorTitle=settings.get('director_title', ''),
        date=today,
        total=total,
        today=dict(today_stats),
        monthStats=dict(month_stats),
        monthStart=month_start,
        trend=rows_to_list(trend),
        byLevel=rows_to_list(by_level),
        behavior={
            'startScore':    start_score,
            'avgScore':      round(score_stats['avg_score'] or start_score, 1),
            'minScore':      score_stats['min_score'] or start_score,
            'deductedCount': score_stats['deducted'] or 0,
            'totalStudents': score_stats['total_students'] or 0,
            'lowScores':     rows_to_list(low_scores),
        },
        frequentAbsent=rows_to_list(frequent_absent),
        alertThreshold=alert_threshold,
    )

# ── REPORT ────────────────────────────────────────────────

def semester_range(year, sem, settings):
    """year is Gregorian, returns (from_iso, to_iso) for a Thai semester."""
    s_start = settings.get(f'sem{sem}_start', '05-16').split('-')
    s_end   = settings.get(f'sem{sem}_end',   '10-15').split('-')
    sm, sd = int(s_start[0]), int(s_start[1])
    em, ed = int(s_end[0]),   int(s_end[1])

    start = datetime.date(year, sm, sd)
    # Semester 2 crosses year boundary if end month < start month
    end_year = year if em >= sm else year + 1
    end = datetime.date(end_year, em, ed)
    return start.isoformat(), end.isoformat()

def build_report_query(level, room, from_date, to_date, scope_where='', scope_params=None):
    scope_params = scope_params or []
    join_cond, join_params = 's.id=a.student_id', []
    if from_date: join_cond += ' AND a.date>=?'; join_params.append(from_date)
    if to_date:   join_cond += ' AND a.date<=?'; join_params.append(to_date)
    where_cond, where_params = '1=1' + scope_where, list(scope_params)
    if level and not scope_where: where_cond += ' AND s.class_level=?'; where_params.append(int(level))
    if room  and not scope_where: where_cond += ' AND s.room=?';        where_params.append(int(room))
    sql = f"""
        SELECT s.id, s.number, s.student_code, s.name, s.class_level, s.room,
            COUNT(a.id) AS total_days,
            SUM(CASE WHEN a.status='present'  THEN 1 ELSE 0 END) AS present,
            SUM(CASE WHEN a.status='absent'   THEN 1 ELSE 0 END) AS absent,
            SUM(CASE WHEN a.status='late'     THEN 1 ELSE 0 END) AS late,
            SUM(CASE WHEN a.status='leave'    THEN 1 ELSE 0 END) AS leave,
            SUM(CASE WHEN a.status='activity' THEN 1 ELSE 0 END) AS activity
        FROM students s
        LEFT JOIN attendance a ON {join_cond}
        WHERE {where_cond}
        GROUP BY s.id
        ORDER BY s.class_level, s.room, s.number, s.name
    """
    return sql, join_params + where_params

@app.get('/api/report')
@login_required
def api_report():
    u = current_user()
    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))

    args = request.args
    sem  = args.get('semester')
    year = args.get('year')
    if sem and year:
        from_d, to_d = semester_range(int(year), int(sem), settings)
    else:
        from_d, to_d = args.get('from'), args.get('to')

    scope_where, scope_params, _, _ = user_class_filter(u, args)
    sql, params = build_report_query(args.get('level'), args.get('room'), from_d, to_d,
                                     scope_where, scope_params)
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()
        # Add behavior scores
        scores = {r['student_id']: r['delta'] for r in con.execute(
            'SELECT student_id, SUM(points) AS delta FROM behavior_logs GROUP BY student_id'
        ).fetchall()}

    out = []
    for r in rows:
        d = dict(r)
        d['score'] = start_score + scores.get(d['id'], 0)
        out.append(d)
    return jsonify(dict(rows=out, range={'from': from_d, 'to': to_d}))

@app.get('/api/report/export')
@login_required
def api_report_export():
    u = current_user()
    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))
    args = request.args
    sem  = args.get('semester'); year = args.get('year')
    if sem and year:
        from_d, to_d = semester_range(int(year), int(sem), settings)
    else:
        from_d, to_d = args.get('from'), args.get('to')

    scope_where, scope_params, _, _ = user_class_filter(u, args)
    sql, params = build_report_query(args.get('level'), args.get('room'), from_d, to_d,
                                     scope_where, scope_params)
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()
        scores = {r['student_id']: r['delta'] for r in con.execute(
            'SELECT student_id, SUM(points) AS delta FROM behavior_logs GROUP BY student_id'
        ).fetchall()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'รายงานการมาเรียน'

    headers = ['เลขที่','รหัสนักเรียน','ชื่อ-นามสกุล','ชั้น','วันที่บันทึก',
               'มา','ขาด','มาสาย','ลา','กิจกรรม','คะแนนความประพฤติ']
    hfill = PatternFill(fill_type='solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            r['number'], r['student_code'], r['name'],
            f"ม.{r['class_level']}/{r['room']}",
            r['total_days'], r['present'], r['absent'], r['late'], r['leave'],
            r['activity'] or 0,
            start_score + scores.get(r['id'], 0)
        ])
    widths = [8, 16, 32, 10, 12, 8, 8, 8, 8, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"attendance_report_{today_iso()}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── PRINT ROLL-CALL SHEET ─────────────────────────────────

@app.get('/api/rollcall/print')
@login_required
def api_rollcall_print():
    u = current_user()
    level = request.args.get('level')
    room  = request.args.get('room')
    if not level or not room:
        return jsonify(error='ต้องระบุชั้น/ห้อง'), 400
    if u['role'] == 'teacher' and u.get('assigned_level'):
        if int(level) != u['assigned_level'] or (u.get('assigned_room') and int(room) != u['assigned_room']):
            return jsonify(error='ไม่มีสิทธิ์เข้าถึงห้องนี้'), 403
    with get_db() as con:
        students = con.execute("""
            SELECT id, number, student_code, name, gender FROM students
            WHERE class_level=? AND room=? ORDER BY number, name
        """, (int(level), int(room))).fetchall()
    return jsonify(students=rows_to_list(students), level=int(level), room=int(room))

# ─────────────────────────────────────────────────────────

# ── HOLIDAYS ──────────────────────────────────────────────

@app.get('/api/holidays')
@login_required
def api_holidays_list():
    year = request.args.get('year')
    sql = 'SELECT * FROM holidays'
    params = []
    if year:
        sql += " WHERE date LIKE ?"; params.append(f'{year}-%')
    sql += ' ORDER BY date'
    with get_db() as con:
        rows = con.execute(sql, params).fetchall()
    return jsonify(rows_to_list(rows))

@app.post('/api/holidays')
@admin_required
def api_holidays_create():
    b = request.get_json() or {}
    date = (b.get('date') or '').strip()
    name = (b.get('name') or '').strip()
    htype = b.get('type', 'holiday')
    if not date or not name:
        return jsonify(success=False, message='กรอกวันที่และชื่อ'), 400
    try:
        with get_db() as con:
            con.execute('INSERT INTO holidays (date, name, type) VALUES (?,?,?)',
                        (date, name, htype))
            audit_log(con, 'create_holiday', 'holiday', None, {'date': date, 'name': name})
    except sqlite3.IntegrityError:
        return jsonify(success=False, message='มีวันหยุดนี้อยู่แล้ว'), 400
    return jsonify(success=True, message='เพิ่มวันหยุดสำเร็จ')

@app.delete('/api/holidays/<date>')
@admin_required
def api_holidays_delete(date):
    with get_db() as con:
        con.execute('DELETE FROM holidays WHERE date=?', (date,))
        audit_log(con, 'delete_holiday', 'holiday', None, {'date': date})
    return jsonify(success=True)

# Seed Thai national holidays for current year
@app.post('/api/holidays/seed')
@admin_required
def api_holidays_seed():
    b = request.get_json() or {}
    year = int(b.get('year', datetime.date.today().year))
    defaults = [
        (f'{year}-01-01', 'วันขึ้นปีใหม่'),
        (f'{year}-04-06', 'วันจักรี'),
        (f'{year}-04-13', 'วันสงกรานต์'),
        (f'{year}-04-14', 'วันสงกรานต์'),
        (f'{year}-04-15', 'วันสงกรานต์'),
        (f'{year}-05-01', 'วันแรงงานแห่งชาติ'),
        (f'{year}-05-04', 'วันฉัตรมงคล'),
        (f'{year}-06-03', 'วันเฉลิมพระชนมพรรษาพระราชินี'),
        (f'{year}-07-28', 'วันเฉลิมพระชนมพรรษา ร.10'),
        (f'{year}-08-12', 'วันแม่แห่งชาติ'),
        (f'{year}-10-13', 'วันคล้ายวันสวรรคต ร.9'),
        (f'{year}-10-23', 'วันปิยมหาราช'),
        (f'{year}-12-05', 'วันพ่อแห่งชาติ'),
        (f'{year}-12-10', 'วันรัฐธรรมนูญ'),
        (f'{year}-12-31', 'วันสิ้นปี'),
    ]
    added = 0
    with get_db() as con:
        for d, n in defaults:
            try:
                con.execute('INSERT INTO holidays (date, name) VALUES (?, ?)', (d, n))
                added += 1
            except sqlite3.IntegrityError: pass
        audit_log(con, 'seed_holidays', 'holiday', None, {'year': year, 'added': added})
    return jsonify(success=True, added=added, message=f'เพิ่ม {added} วันหยุดราชการปี {year}')

# ── BACKUP / RESTORE ──────────────────────────────────────

@app.get('/api/backup')
@admin_required
def api_backup():
    """Download zip of DB + photos."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(DB_PATH, 'school.db')
        # Photos
        for f in os.listdir(PHOTOS_DIR):
            zf.write(os.path.join(PHOTOS_DIR, f), f'photos/{f}')
        # Assets (logo)
        for f in os.listdir(ASSETS_DIR):
            zf.write(os.path.join(ASSETS_DIR, f), f'assets/{f}')
        manifest = {'created_at': datetime.datetime.now().isoformat(),
                    'created_by': current_user()['full_name']}
        zf.writestr('manifest.json', json.dumps(manifest, ensure_ascii=False, indent=2))
    buf.seek(0)
    with get_db() as con:
        audit_log(con, 'backup', 'system', None, None)
    return send_file(buf, as_attachment=True,
                     download_name=f'attendance_backup_{today_iso()}.zip',
                     mimetype='application/zip')

@app.post('/api/restore')
@admin_required
def api_restore():
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.zip'):
        return jsonify(success=False, message='ต้องเป็นไฟล์ .zip'), 400
    try:
        zbuf = io.BytesIO(f.read())
        with zipfile.ZipFile(zbuf, 'r') as zf:
            names = zf.namelist()
            if 'school.db' not in names:
                return jsonify(success=False, message='ไฟล์ backup ไม่ถูกต้อง (ไม่พบ school.db)'), 400
            # Restore DB
            with zf.open('school.db') as src, open(DB_PATH, 'wb') as dst:
                shutil.copyfileobj(src, dst)
            # Restore photos
            for n in names:
                if n.startswith('photos/') and not n.endswith('/'):
                    target = os.path.join(PHOTOS_DIR, os.path.basename(n))
                    with zf.open(n) as src, open(target, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
                elif n.startswith('assets/') and not n.endswith('/'):
                    target = os.path.join(ASSETS_DIR, os.path.basename(n))
                    with zf.open(n) as src, open(target, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
    except Exception as e:
        return jsonify(success=False, message=f'กู้คืนไม่สำเร็จ: {e}'), 500
    return jsonify(success=True, message='กู้คืนสำเร็จ — กรุณา login ใหม่')

# ── ADMIN: BACKUP DATABASE ────────────────────────────────

@app.get('/api/admin/backup')
@admin_required
def api_admin_backup():
    """ดาวน์โหลดไฟล์ฐานข้อมูล SQLite (สำหรับเก็บสำรอง)"""
    import io, tempfile
    # ใช้ sqlite backup API เพื่อ snapshot ที่ consistent (ไม่กระทบ WAL)
    try:
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
            tmp_path = tmp.name
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(tmp_path)
        with dst:
            src.backup(dst)
        src.close()
        dst.close()

        with open(tmp_path, 'rb') as f:
            data = f.read()
        os.unlink(tmp_path)

        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'school_backup_{ts}.db'
        audit_log_oneshot('download_backup', 'system', None, {'file': filename, 'bytes': len(data)})

        return send_file(
            io.BytesIO(data),
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify(success=False, message=f'สร้าง backup ไม่สำเร็จ: {e}'), 500

@app.get('/api/admin/backup/info')
@admin_required
def api_admin_backup_info():
    """ข้อมูล DB เบื้องต้น: ขนาดไฟล์, จำนวน records"""
    try:
        size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
        with get_db() as con:
            counts = {
                'students':      con.execute('SELECT COUNT(*) AS n FROM students').fetchone()['n'],
                'attendance':    con.execute('SELECT COUNT(*) AS n FROM attendance').fetchone()['n'],
                'behavior_logs': con.execute('SELECT COUNT(*) AS n FROM behavior_logs').fetchone()['n'],
                'users':         con.execute('SELECT COUNT(*) AS n FROM users').fetchone()['n'],
            }
        return jsonify(size_bytes=size, size_mb=round(size/1024/1024, 2), counts=counts)
    except Exception as e:
        return jsonify(error=str(e)), 500

# helper สำหรับ audit log ใน endpoint ที่ไม่มี context db อยู่แล้ว
def audit_log_oneshot(action, entity, entity_id, payload):
    try:
        with get_db() as con:
            audit_log(con, action, entity, entity_id, payload)
    except Exception:
        pass

# ── ADMIN: RESET DATA ─────────────────────────────────────

@app.post('/api/admin/reset')
@admin_required
def api_admin_reset():
    b = request.get_json() or {}
    reset_behavior = b.get('reset_behavior', False)
    reset_attendance = b.get('reset_attendance', False)
    before = b.get('before')  # optional ISO date

    deleted = {}
    with get_db() as con:
        if reset_behavior:
            if before:
                cur = con.execute('DELETE FROM behavior_logs WHERE date < ?', (before,))
            else:
                cur = con.execute('DELETE FROM behavior_logs')
            deleted['behavior_logs'] = cur.rowcount
        if reset_attendance:
            if before:
                cur = con.execute('DELETE FROM attendance WHERE date < ?', (before,))
            else:
                cur = con.execute('DELETE FROM attendance')
            deleted['attendance'] = cur.rowcount
        audit_log(con, 'reset_data', 'system', None,
                  {'before': before, 'deleted': deleted,
                   'reset_behavior': reset_behavior, 'reset_attendance': reset_attendance})

    return jsonify(success=True, deleted=deleted,
                   message='รีเซ็ตข้อมูลสำเร็จ ' + json.dumps(deleted))

# ── AUDIT LOG ─────────────────────────────────────────────

@app.get('/api/audit')
@admin_required
def api_audit_list():
    limit = int(request.args.get('limit', 200))
    action = request.args.get('action')
    user_id = request.args.get('user_id')
    cond, params = '1=1', []
    if action: cond += ' AND action=?'; params.append(action)
    if user_id: cond += ' AND user_id=?'; params.append(int(user_id))
    with get_db() as con:
        rows = con.execute(
            f'SELECT * FROM audit_logs WHERE {cond} ORDER BY id DESC LIMIT ?',
            params + [limit]
        ).fetchall()
    return jsonify(rows_to_list(rows))

# ── GLOBAL STUDENT SEARCH ─────────────────────────────────

@app.get('/api/students/search')
@login_required
def api_students_search():
    q = (request.args.get('q') or '').strip()
    if not q or len(q) < 1:
        return jsonify([])
    u = current_user()
    where, params, _, _ = user_class_filter(u)
    pattern = f'%{q}%'
    sql = f"""
        SELECT id, number, student_code, name, class_level, room, photo
        FROM students s
        WHERE (name LIKE ? OR student_code LIKE ? OR CAST(number AS TEXT) LIKE ?) {where}
        ORDER BY class_level, room, number, name LIMIT 30
    """
    with get_db() as con:
        rows = con.execute(sql, [pattern, pattern, pattern] + params).fetchall()
    return jsonify(rows_to_list(rows))

# ── HOME VISITS (เยี่ยมบ้าน) ───────────────────────────────

@app.get('/api/home-visits')
@login_required
def api_home_visits_list():
    """รายชื่อนักเรียนในห้อง + สถานะเยี่ยมบ้าน — Query: ?level=1&room=1"""
    u = current_user()
    level = request.args.get('level')
    room = request.args.get('room')
    where, params, _, _ = user_class_filter(u, {'level': level, 'room': room})
    with get_db() as con:
        rows = con.execute(f"""
            SELECT s.id, s.number, s.student_code, s.name, s.class_level, s.room, s.photo,
                   hv.id AS visit_id, hv.visited, hv.visit_date,
                   hv.photos, hv.submitted, hv.confirmed
            FROM students s
            LEFT JOIN home_visits hv ON hv.student_id = s.id
            WHERE 1=1 {where}
            ORDER BY s.number, s.name
        """, params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        try: d['photo_count'] = len(json.loads(d.get('photos') or '[]'))
        except Exception: d['photo_count'] = 0
        d.pop('photos', None)
        result.append(d)
    return jsonify(result)

@app.get('/api/home-visits/export')
@login_required
def api_home_visits_export():
    """Export ข้อมูลเยี่ยมบ้านทั้งห้องเป็น Excel — Query: ?level=1&room=1"""
    u = current_user()
    level = request.args.get('level')
    room = request.args.get('room')
    where, params, _, _ = user_class_filter(u, {'level': level, 'room': room})
    with get_db() as con:
        rows = con.execute(f"""
            SELECT s.number, s.student_code, s.name, s.class_level, s.room,
                   s.birthdate, s.national_id,
                   hv.visited, hv.visit_date, hv.data
            FROM students s
            LEFT JOIN home_visits hv ON hv.student_id = s.id
            WHERE 1=1 {where}
            ORDER BY s.class_level, s.room, s.number, s.name
        """, params).fetchall()

    # คอลัมน์ตาม schema (key, หัวตาราง)
    fields = [
        ('nickname','ชื่อเล่น'), ('nationality','สัญชาติ'), ('ethnicity','เชื้อชาติ'),
        ('religion','ศาสนา'), ('age','อายุ'),
        ('prev_school','โรงเรียนเดิม'), ('prev_gpa','เกรดเฉลี่ย'),
        ('weight','น้ำหนัก'), ('height','ส่วนสูง'), ('blood','หมู่เลือด'), ('disease','โรคประจำตัว'),
        ('address','ที่อยู่'),
        ('q13_house','สถานะที่อยู่อาศัย'), ('q14_house_type','ลักษณะบ้าน'),
        ('q15_burden','ภาระพึ่งพิง'), ('q16_safety','ความปลอดภัยบ้าน'),
        ('q17_living_with','พักอาศัยอยู่กับ'), ('q18_travel','การเดินทาง'),
        ('q18_cost','ค่าเดินทาง/เดือน'), ('q19_distance','ระยะทาง'), ('q20_time','เวลาเดินทาง'),
        ('q21_route_safety','ความปลอดภัยระหว่างทาง'), ('q22_electric','แหล่งไฟฟ้า'),
        ('q23_water','แหล่งน้ำดื่ม'), ('q24_disaster','ความเสี่ยงภัยพิบัติ'),
        ('q25_friend','เพื่อนสนิท'),
        ('q27_need','สภาพที่น่าเป็นห่วง'), ('q27_summary','ข้อสรุปเยี่ยมบ้าน'),
    ]

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'ข้อมูลเยี่ยมบ้าน'
    headers = ['เลขที่','รหัสนักเรียน','ชื่อ-สกุล','ชั้น','ห้อง','วันเกิด','เลขบัตรประชาชน',
               'เยี่ยมแล้ว','วันที่เยี่ยม'] + [h for _, h in fields]
    hfill = PatternFill(fill_type='solid', fgColor='1A5276')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal='center', wrap_text=True)

    for r in rows:
        try: data = json.loads(r['data'] or '{}')
        except Exception: data = {}
        line = [
            r['number'], r['student_code'], r['name'], r['class_level'], r['room'],
            r['birthdate'], r['national_id'],
            'เยี่ยมแล้ว' if r['visited'] else 'ยังไม่เยี่ยม', r['visit_date'],
        ]
        for key, _ in fields:
            v = data.get(key, '')
            if isinstance(v, list): v = ', '.join(v)
            line.append(v)
        ws.append(line)

    widths = [7, 14, 26, 6, 6, 14, 18, 12, 14] + [18]*len(fields)
    for col, w in zip(ws.columns, widths):
        ws.column_dimensions[col[0].column_letter].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    lv = level or 'all'; rm = room or 'all'
    return send_file(buf, as_attachment=True,
                     download_name=f'home_visit_m{lv}_{rm}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/api/home-visits/<int:sid>')
@login_required
def api_home_visit_get(sid):
    """ข้อมูลเยี่ยมบ้านของนักเรียน 1 คน (สร้าง record ว่างถ้ายังไม่มี)"""
    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student:
            return jsonify(error='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(error='ไม่มีสิทธิ์'), 403
        hv = con.execute('SELECT * FROM home_visits WHERE student_id=?', (sid,)).fetchone()
    student_d = {
        'id': student['id'], 'number': student['number'], 'name': student['name'],
        'student_code': student['student_code'], 'class_level': student['class_level'],
        'room': student['room'], 'photo': student['photo'],
        'birthdate': student['birthdate'], 'national_id': student['national_id'],
        'parent_name': student['parent_name'], 'parent_relation': student['parent_relation'],
        'parent_phone': student['parent_phone'], 'address': student['address'],
    }
    if hv:
        d = dict(hv)
        try: d['data'] = json.loads(d.get('data') or '{}')
        except Exception: d['data'] = {}
        try: d['photos'] = json.loads(d.get('photos') or '[]')
        except Exception: d['photos'] = []
    else:
        d = {'student_id': sid, 'visited': 0, 'visit_date': None, 'data': {}, 'photos': []}
    d['student'] = student_d
    return jsonify(d)

@app.post('/api/home-visits/<int:sid>')
@login_required
def api_home_visit_save(sid):
    """บันทึก/อัพเดทข้อมูลเยี่ยมบ้าน"""
    u = current_user()
    b = request.get_json() or {}
    visited    = 1 if b.get('visited') else 0
    visit_date = (b.get('visit_date') or '').strip() or None
    data       = b.get('data') or {}
    data_json  = json.dumps(data, ensure_ascii=False)

    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student:
            return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        existing = con.execute('SELECT id FROM home_visits WHERE student_id=?', (sid,)).fetchone()
        if existing:
            con.execute("""
                UPDATE home_visits SET visited=?, visit_date=?, data=?,
                       recorded_by=?, updated_at=datetime('now','localtime')
                WHERE student_id=?
            """, (visited, visit_date, data_json, u['id'], sid))
        else:
            con.execute("""
                INSERT INTO home_visits (student_id, visited, visit_date, data, photos, recorded_by)
                VALUES (?,?,?,?,?,?)
            """, (sid, visited, visit_date, data_json, '[]', u['id']))
        audit_log(con, 'save_home_visit', 'student', sid, {'visited': visited})
    return jsonify(success=True, message='บันทึกข้อมูลเยี่ยมบ้านแล้ว')

@app.post('/api/home-visits/<int:sid>/fill-link')
@login_required
def api_home_visit_gen_link(sid):
    """สร้าง/ดึงรหัสลิงก์ให้นักเรียนกรอกเอง"""
    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        hv = con.execute('SELECT id, fill_code FROM home_visits WHERE student_id=?', (sid,)).fetchone()
        code = hv['fill_code'] if hv and hv['fill_code'] else None
        if not code:
            for _ in range(5):
                code = secrets.token_urlsafe(8)
                try:
                    if hv:
                        con.execute('UPDATE home_visits SET fill_code=? WHERE student_id=?', (code, sid))
                    else:
                        con.execute('INSERT INTO home_visits (student_id, fill_code, photos) VALUES (?,?,?)',
                                    (sid, code, '[]'))
                    break
                except sqlite3.IntegrityError:
                    code = None; continue
        audit_log(con, 'gen_hv_fill_link', 'student', sid, None)
    return jsonify(success=True, code=code, url=f'/hv-fill.html?code={code}')

@app.post('/api/home-visits/<int:sid>/confirm')
@login_required
def api_home_visit_confirm(sid):
    """ครูยืนยันข้อมูลที่นักเรียนกรอก"""
    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        con.execute("UPDATE home_visits SET confirmed=1, visited=1, recorded_by=?, "
                    "updated_at=datetime('now','localtime') WHERE student_id=?", (u['id'], sid))
        audit_log(con, 'confirm_home_visit', 'student', sid, None)
    return jsonify(success=True, message='ยืนยันข้อมูลแล้ว')

# ----- PUBLIC: นักเรียนกรอกเองด้วย fill_code (ไม่ต้องล็อกอิน) -----

def _hv_by_code(con, code):
    return con.execute("""
        SELECT hv.*, s.name AS student_name, s.class_level, s.room, s.number,
               s.student_code, s.birthdate, s.address
        FROM home_visits hv JOIN students s ON s.id=hv.student_id
        WHERE hv.fill_code=?
    """, (code,)).fetchone()

@app.get('/api/hv-fill/<code>')
def api_hv_fill_get(code):
    """ดึงข้อมูลฟอร์มสำหรับนักเรียนกรอก (public)"""
    with get_db() as con:
        hv = _hv_by_code(con, code)
    if not hv:
        return jsonify(error='ลิงก์ไม่ถูกต้องหรือหมดอายุ'), 404
    d = dict(hv)
    try: d['data'] = json.loads(d.get('data') or '{}')
    except Exception: d['data'] = {}
    try: d['photos'] = json.loads(d.get('photos') or '[]')
    except Exception: d['photos'] = []
    # ส่งเฉพาะที่จำเป็น (ไม่หลุดข้อมูลอ่อนไหว)
    return jsonify(
        student={'name': d['student_name'], 'class_level': d['class_level'],
                 'room': d['room'], 'number': d['number'],
                 'birthdate': d['birthdate'], 'address': d['address']},
        data=d['data'], photos=d['photos'],
        submitted=d['submitted'], confirmed=d['confirmed'],
    )

@app.post('/api/hv-fill/<code>')
def api_hv_fill_save(code):
    """นักเรียนกรอก+ส่ง (public) — set submitted=1 รอครูยืนยัน"""
    b = request.get_json() or {}
    data = b.get('data') or {}
    visit_date = (b.get('visit_date') or '').strip() or None
    with get_db() as con:
        hv = _hv_by_code(con, code)
        if not hv:
            return jsonify(success=False, message='ลิงก์ไม่ถูกต้อง'), 404
        if hv['confirmed']:
            return jsonify(success=False, message='ครูยืนยันข้อมูลแล้ว ไม่สามารถแก้ไขได้'), 403
        con.execute("""
            UPDATE home_visits SET data=?, visit_date=COALESCE(?,visit_date), submitted=1,
                   updated_at=datetime('now','localtime')
            WHERE fill_code=?
        """, (json.dumps(data, ensure_ascii=False), visit_date, code))
    return jsonify(success=True, message='ส่งข้อมูลเรียบร้อย ขอบคุณค่ะ/ครับ')

@app.post('/api/hv-fill/<code>/photo')
def api_hv_fill_photo(code):
    """นักเรียนอัพรูป (public)"""
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    if not f.filename: return jsonify(success=False, message='ไม่พบไฟล์'), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
        return jsonify(success=False, message='ต้องเป็นรูปภาพ'), 400
    with get_db() as con:
        hv = _hv_by_code(con, code)
        if not hv: return jsonify(success=False, message='ลิงก์ไม่ถูกต้อง'), 404
        if hv['confirmed']:
            return jsonify(success=False, message='ยืนยันแล้ว แก้ไขไม่ได้'), 403
        filename = f'hv_{hv["student_id"]}_{secrets.token_hex(4)}{ext}'
        f.save(os.path.join(PHOTOS_DIR, filename))
        try: photos = json.loads(hv['photos'] or '[]')
        except Exception: photos = []
        photos.append(filename)
        con.execute('UPDATE home_visits SET photos=? WHERE fill_code=?',
                    (json.dumps(photos, ensure_ascii=False), code))
    return jsonify(success=True, photo=filename)

@app.post('/api/home-visits/<int:sid>/photo')
@login_required
def api_home_visit_photo(sid):
    """อัพโหลดรูปเยี่ยมบ้าน (เพิ่มเข้า list)"""
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    if not f.filename: return jsonify(success=False, message='ไม่พบไฟล์'), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
        return jsonify(success=False, message='ต้องเป็นรูปภาพ jpg/png/webp'), 400

    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403

        # filename ไม่ซ้ำ
        filename = f'hv_{sid}_{secrets.token_hex(4)}{ext}'
        f.save(os.path.join(PHOTOS_DIR, filename))

        hv = con.execute('SELECT id, photos FROM home_visits WHERE student_id=?', (sid,)).fetchone()
        if hv:
            try: photos = json.loads(hv['photos'] or '[]')
            except Exception: photos = []
            photos.append(filename)
            con.execute('UPDATE home_visits SET photos=?, updated_at=datetime(\'now\',\'localtime\') WHERE student_id=?',
                        (json.dumps(photos, ensure_ascii=False), sid))
        else:
            con.execute("""
                INSERT INTO home_visits (student_id, visited, photos, recorded_by)
                VALUES (?, 0, ?, ?)
            """, (sid, json.dumps([filename], ensure_ascii=False), u['id']))
        audit_log(con, 'upload_hv_photo', 'student', sid, {'filename': filename})
    return jsonify(success=True, message='อัพโหลดรูปสำเร็จ', photo=filename)

@app.delete('/api/home-visits/<int:sid>/photo')
@login_required
def api_home_visit_photo_delete(sid):
    """ลบรูปเยี่ยมบ้าน 1 รูป — Body: { filename }"""
    u = current_user()
    b = request.get_json() or {}
    filename = b.get('filename')
    if not filename: return jsonify(success=False, message='ไม่พบชื่อไฟล์'), 400
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        hv = con.execute('SELECT photos FROM home_visits WHERE student_id=?', (sid,)).fetchone()
        if hv:
            try: photos = json.loads(hv['photos'] or '[]')
            except Exception: photos = []
            if filename in photos:
                photos.remove(filename)
                con.execute('UPDATE home_visits SET photos=? WHERE student_id=?',
                            (json.dumps(photos, ensure_ascii=False), sid))
                try: os.remove(os.path.join(PHOTOS_DIR, filename))
                except OSError: pass
    return jsonify(success=True, message='ลบรูปแล้ว')

# ── PHOTO UPLOAD ──────────────────────────────────────────

@app.delete('/api/students/<int:sid>/photo')
@login_required
def api_student_photo_delete(sid):
    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        if student['photo']:
            try:
                os.remove(os.path.join(PHOTOS_DIR, student['photo']))
            except OSError:
                pass
        con.execute('UPDATE students SET photo=NULL WHERE id=?', (sid,))
        audit_log(con, 'delete_photo', 'student', sid, None)
    return jsonify(success=True, message='ลบรูปแล้ว')

@app.post('/api/students/<int:sid>/photo')
@login_required
def api_student_photo(sid):
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    if not f.filename: return jsonify(success=False, message='ไม่พบไฟล์'), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
        return jsonify(success=False, message='ต้องเป็นรูปภาพ jpg/png/webp'), 400

    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403

        filename = f'student_{sid}{ext}'
        f.save(os.path.join(PHOTOS_DIR, filename))
        con.execute('UPDATE students SET photo=? WHERE id=?', (filename, sid))
        audit_log(con, 'upload_photo', 'student', sid, {'filename': filename})

    return jsonify(success=True, message='อัพโหลดรูปสำเร็จ', photo=filename)

# ── PARENT PORTAL ─────────────────────────────────────────

@app.post('/api/students/<int:sid>/parent-code')
@login_required
def api_gen_parent_code(sid):
    """Generate or regenerate parent access code for a student."""
    u = current_user()
    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE id=?', (sid,)).fetchone()
        if not student: return jsonify(success=False, message='ไม่พบนักเรียน'), 404
        if not can_access_student(u, student):
            return jsonify(success=False, message='ไม่มีสิทธิ์'), 403
        # Try a few times in case of collision
        for _ in range(5):
            code = generate_parent_code()
            try:
                con.execute('UPDATE students SET parent_code=? WHERE id=?', (code, sid))
                audit_log(con, 'gen_parent_code', 'student', sid, None)
                return jsonify(success=True, code=code,
                               url=f'/parent.html?code={code}')
            except sqlite3.IntegrityError: continue
    return jsonify(success=False, message='สร้างรหัสไม่สำเร็จ'), 500

@app.post('/api/students/parent-codes/bulk')
@login_required
def api_gen_parent_codes_bulk():
    """Generate codes for all students in a class who don't have one."""
    u = current_user()
    b = request.get_json() or {}
    level = b.get('level')
    room = b.get('room')
    where, params, _, _ = user_class_filter(u, {'level': level, 'room': room})
    with get_db() as con:
        rows = con.execute(
            f'SELECT id FROM students s WHERE (parent_code IS NULL OR parent_code = "") {where}',
            params
        ).fetchall()
        count = 0
        for r in rows:
            for _ in range(5):
                code = generate_parent_code()
                try:
                    con.execute('UPDATE students SET parent_code=? WHERE id=?', (code, r['id']))
                    count += 1
                    break
                except sqlite3.IntegrityError: continue
        audit_log(con, 'bulk_gen_parent_codes', 'student', None, {'count': count})
    return jsonify(success=True, count=count, message=f'สร้างรหัสผู้ปกครองให้ {count} คน')

@app.get('/api/parent/<code>')
def api_parent_view(code):
    """Public endpoint for parents (no auth) — view child's info."""
    code = code.upper()
    settings = get_settings()
    start_score = int(settings.get('start_score', '100'))
    today_d = datetime.date.today()
    month_start = today_d.replace(day=1).isoformat()

    with get_db() as con:
        student = con.execute('SELECT * FROM students WHERE parent_code=?', (code,)).fetchone()
        if not student: return jsonify(error='ไม่พบรหัสนี้'), 404
        sid = student['id']

        att_all = con.execute("""
            SELECT date, status, note FROM attendance
            WHERE student_id=? ORDER BY date DESC LIMIT 60
        """, (sid,)).fetchall()

        stats = con.execute("""
            SELECT
              SUM(status='present')  AS present,
              SUM(status='absent')   AS absent,
              SUM(status='late')     AS late,
              SUM(status='leave')    AS leave,
              SUM(status='activity') AS activity,
              COUNT(*) AS total
            FROM attendance WHERE student_id=? AND date >= ?
        """, (sid, month_start)).fetchone()

        bhv = con.execute("""
            SELECT date, points, reason, source FROM behavior_logs
            WHERE student_id=? ORDER BY date DESC, id DESC LIMIT 30
        """, (sid,)).fetchall()
        delta = con.execute('SELECT COALESCE(SUM(points),0) AS d FROM behavior_logs WHERE student_id=?',
                            (sid,)).fetchone()['d']

    return jsonify(
        student={
            'name': student['name'], 'number': student['number'],
            'class_level': student['class_level'], 'room': student['room'],
            'student_code': student['student_code'], 'photo': student['photo']
        },
        attendance={'recent': rows_to_list(att_all), 'month': dict(stats)},
        behavior={'score': start_score + delta, 'start_score': start_score,
                  'logs': rows_to_list(bhv)},
        school_name=settings.get('school_name'),
        school_logo=settings.get('school_logo')
    )

# ── CHARTS DATA ───────────────────────────────────────────

@app.get('/api/charts/weekly')
@login_required
def api_chart_weekly():
    """Attendance trend over last 14 days."""
    u = current_user()
    where, params, _, _ = user_class_filter(u)
    end = datetime.date.today()
    start = end - datetime.timedelta(days=13)

    with get_db() as con:
        rows = con.execute(f"""
            SELECT a.date,
                   SUM(status='present')  AS present,
                   SUM(status='absent')   AS absent,
                   SUM(status='late')     AS late,
                   SUM(status='leave')    AS leave,
                   SUM(status='activity') AS activity
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE a.date >= ? AND a.date <= ? {where}
            GROUP BY a.date ORDER BY a.date
        """, [start.isoformat(), end.isoformat()] + params).fetchall()
    return jsonify(rows_to_list(rows))

@app.get('/api/charts/class-summary')
@login_required
def api_chart_class_summary():
    """Total absences per class (this month)."""
    u = current_user()
    where, params, _, _ = user_class_filter(u)
    today_d = datetime.date.today()
    month_start = today_d.replace(day=1).isoformat()

    with get_db() as con:
        rows = con.execute(f"""
            SELECT s.class_level, s.room,
                   SUM(a.status='absent') AS absent,
                   SUM(a.status='late') AS late
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE a.date >= ? {where}
            GROUP BY s.class_level, s.room
            ORDER BY s.class_level, s.room
        """, [month_start] + params).fetchall()
    return jsonify(rows_to_list(rows))

# ── LOGO UPLOAD ───────────────────────────────────────────

@app.post('/api/settings/logo')
@admin_required
def api_upload_logo():
    if 'file' not in request.files:
        return jsonify(success=False, message='ไม่พบไฟล์'), 400
    f = request.files['file']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.png', '.jpg', '.jpeg', '.svg', '.webp'):
        return jsonify(success=False, message='รองรับ png/jpg/svg/webp'), 400
    filename = f'school_logo{ext}'
    # Remove old logos with different extensions
    for old_ext in ('.png','.jpg','.jpeg','.svg','.webp'):
        old = os.path.join(ASSETS_DIR, f'school_logo{old_ext}')
        if os.path.exists(old) and old_ext != ext: os.remove(old)
    f.save(os.path.join(ASSETS_DIR, filename))
    with get_db() as con:
        con.execute('UPDATE settings SET value=? WHERE key=?', (f'/assets/{filename}', 'school_logo'))
        audit_log(con, 'upload_logo', 'settings', None, {'filename': filename})
    return jsonify(success=True, message='อัพโหลดโลโก้สำเร็จ', path=f'/assets/{filename}')

def cli_reset_admin():
    """Reset admin password to admin123."""
    new_pw = 'admin123'
    with get_db() as con:
        u = con.execute('SELECT id FROM users WHERE username=?', ('admin',)).fetchone()
        if u:
            con.execute('UPDATE users SET password_hash=? WHERE username=?',
                        (hash_password(new_pw), 'admin'))
            print(f"✓ รีเซ็ตรหัสผ่าน admin เรียบร้อย — รหัสใหม่: {new_pw}")
        else:
            con.execute(
                'INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)',
                ('admin', hash_password(new_pw), 'ผู้ดูแลระบบ', 'admin')
            )
            print(f"✓ สร้างบัญชี admin ใหม่ — รหัส: {new_pw}")
    print("⚠️  กรุณาเปลี่ยนรหัสผ่านทันทีหลัง login!")

def cli_reset_user(username, password=None):
    """Reset any user's password."""
    if not password:
        password = secrets.token_urlsafe(8)
    with get_db() as con:
        u = con.execute('SELECT id, full_name, role FROM users WHERE username=?',
                        (username,)).fetchone()
        if not u:
            print(f"✗ ไม่พบผู้ใช้: {username}")
            return
        con.execute('UPDATE users SET password_hash=? WHERE username=?',
                    (hash_password(password), username))
        print(f"✓ รีเซ็ตรหัสผ่าน '{username}' ({u['full_name']}, {u['role']}) เรียบร้อย")
        print(f"  รหัสใหม่: {password}")

def cli_list_users():
    with get_db() as con:
        rows = con.execute(
            'SELECT username, full_name, role, assigned_level, assigned_room FROM users ORDER BY role, username'
        ).fetchall()
    print(f"\nรายชื่อผู้ใช้ ({len(rows)} คน):")
    for r in rows:
        cls = ''
        if r['assigned_level']:
            cls = f" → ม.{r['assigned_level']}" + (f"/{r['assigned_room']}" if r['assigned_room'] else '')
        print(f"  [{r['role']:7}] {r['username']:15} {r['full_name']}{cls}")
    print()

if __name__ == '__main__':
    import sys
    args = sys.argv[1:]

    # CLI tools
    if args:
        if args[0] == '--reset-admin':
            cli_reset_admin(); sys.exit(0)
        elif args[0] == '--reset-user':
            if len(args) < 2:
                print("Usage: python app.py --reset-user <username> [<new-password>]")
                sys.exit(1)
            cli_reset_user(args[1], args[2] if len(args) > 2 else None)
            sys.exit(0)
        elif args[0] == '--list-users':
            cli_list_users(); sys.exit(0)
        elif args[0] in ('--help', '-h'):
            print("""
ระบบเช็คชื่อนักเรียน — โรงเรียนตะกั่วทุ่งงานทวีวิทยาคม

Usage:
  python app.py                              เริ่มเซิร์ฟเวอร์
  python app.py --reset-admin                รีเซ็ตรหัสผ่าน admin เป็น admin123
  python app.py --reset-user <user> [<pw>]   รีเซ็ตรหัสผ่านผู้ใช้ (ไม่ใส่ pw จะสุ่มให้)
  python app.py --list-users                 แสดงรายชื่อผู้ใช้ทั้งหมด
  python app.py --help                       แสดงข้อความนี้

Env:
  PORT=8080                                  พอร์ตเซิร์ฟเวอร์
""")
            sys.exit(0)
        else:
            print(f"Unknown argument: {args[0]}")
            print("ดู: python app.py --help")
            sys.exit(1)

    print('\n========================================')
    print('  ระบบเช็คชื่อนักเรียน')
    print('  โรงเรียนตะกั่วทุ่งงานทวีวิทยาคม')
    print('========================================')
    port = int(os.environ.get('PORT', 8080))
    print(f'  http://localhost:{port}')
    print('  ลืมรหัสผ่าน admin? → python app.py --reset-admin')
    print('========================================\n')
    app.run(host='0.0.0.0', port=port, debug=False)
